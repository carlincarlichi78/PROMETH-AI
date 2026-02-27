"""SFCE — Exportador universal (CSV/Excel)."""

import csv
import io
from datetime import date
from pathlib import Path
from typing import Any

from scripts.core.logger import crear_logger

logger = crear_logger("exportador")


class Exportador:
    """Exporta datos contables a CSV y Excel."""

    def exportar_libro_diario_csv(self, asientos: list[dict],
                                    ruta: str | Path) -> Path:
        """Exporta libro diario a CSV.

        asientos: lista de {"fecha", "numero", "concepto", "partidas": [
            {"subcuenta", "debe", "haber", "concepto"}
        ]}
        """
        ruta = Path(ruta)
        with open(ruta, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow([
                "Asiento", "Fecha", "Subcuenta", "Debe", "Haber", "Concepto"
            ])
            for asiento in asientos:
                for partida in asiento.get("partidas", []):
                    writer.writerow([
                        asiento.get("numero", ""),
                        asiento.get("fecha", ""),
                        partida.get("subcuenta", ""),
                        self._formato_decimal(partida.get("debe", 0)),
                        self._formato_decimal(partida.get("haber", 0)),
                        partida.get("concepto", asiento.get("concepto", "")),
                    ])
        logger.info(f"Exportado libro diario: {ruta}")
        return ruta

    def exportar_facturas_csv(self, facturas: list[dict],
                               ruta: str | Path, tipo: str = "recibidas") -> Path:
        """Exporta listado de facturas a CSV.

        facturas: lista de {"numero", "fecha", "cif", "nombre", "base",
                            "iva", "irpf", "total", "pagada"}
        """
        ruta = Path(ruta)
        with open(ruta, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            if tipo == "recibidas":
                writer.writerow([
                    "Numero", "Fecha", "CIF Emisor", "Nombre Emisor",
                    "Base Imponible", "IVA", "IRPF", "Total", "Pagada"
                ])
            else:
                writer.writerow([
                    "Numero", "Fecha", "CIF Receptor", "Nombre Receptor",
                    "Base Imponible", "IVA", "IRPF", "Total", "Cobrada"
                ])
            for fac in facturas:
                writer.writerow([
                    fac.get("numero", ""),
                    fac.get("fecha", ""),
                    fac.get("cif", ""),
                    fac.get("nombre", ""),
                    self._formato_decimal(fac.get("base", 0)),
                    self._formato_decimal(fac.get("iva", 0)),
                    self._formato_decimal(fac.get("irpf", 0)),
                    self._formato_decimal(fac.get("total", 0)),
                    "Si" if fac.get("pagada") else "No",
                ])
        logger.info(f"Exportado facturas {tipo}: {ruta}")
        return ruta

    def exportar_excel_multihoja(self, datos: dict[str, list[dict]],
                                  ruta: str | Path) -> Path:
        """Exporta multiples hojas a Excel.

        datos: {"nombre_hoja": [{"col1": val1, "col2": val2}, ...]}
        """
        import openpyxl
        ruta = Path(ruta)
        wb = openpyxl.Workbook()

        primera = True
        for nombre_hoja, filas in datos.items():
            if primera:
                ws = wb.active
                ws.title = nombre_hoja
                primera = False
            else:
                ws = wb.create_sheet(nombre_hoja)

            if not filas:
                continue

            # Cabeceras
            cabeceras = list(filas[0].keys())
            for col, cab in enumerate(cabeceras, 1):
                ws.cell(row=1, column=col, value=cab)

            # Datos
            for row_idx, fila in enumerate(filas, 2):
                for col, cab in enumerate(cabeceras, 1):
                    ws.cell(row=row_idx, column=col, value=fila.get(cab))

        wb.save(ruta)
        logger.info(f"Exportado Excel: {ruta}")
        return ruta

    def exportar_libro_diario_excel(self, asientos: list[dict],
                                     ruta: str | Path) -> Path:
        """Exporta libro diario a Excel con formato."""
        filas = []
        for asiento in asientos:
            for partida in asiento.get("partidas", []):
                filas.append({
                    "Asiento": asiento.get("numero", ""),
                    "Fecha": asiento.get("fecha", ""),
                    "Subcuenta": partida.get("subcuenta", ""),
                    "Debe": partida.get("debe", 0),
                    "Haber": partida.get("haber", 0),
                    "Concepto": partida.get("concepto",
                                            asiento.get("concepto", "")),
                })
        return self.exportar_excel_multihoja({"Libro Diario": filas}, ruta)

    # --- Helper ---
    def _formato_decimal(self, valor) -> str:
        """Formatea numero a 2 decimales."""
        if valor is None:
            return "0.00"
        try:
            return f"{float(valor):.2f}"
        except (ValueError, TypeError):
            return "0.00"
