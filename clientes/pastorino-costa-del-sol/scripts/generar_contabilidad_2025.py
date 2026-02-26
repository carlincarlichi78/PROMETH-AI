"""
Script para generar el libro contable 2025 de Pastorino Costa del Sol S.L.
Datos extraidos de las facturas PDF en PROYECTO PASTORINO/dashboard/documentos/gestor/
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os
import shutil
from datetime import date

# ============================================================
# DATOS EXTRAIDOS DE LAS FACTURAS
# ============================================================

# --- FACTURAS EMITIDAS (Pastorino vende limones a Malaga Natural) ---
facturas_emitidas = [
    {
        "fecha": "09/08/2025", "trimestre": "T3", "num_factura": "INV/2025/00001",
        "cliente": "MALAGA NATURAL 2012 S.L.", "cif_cliente": "B93159044",
        "concepto": "Venta limon Genova - Container SUDU8020260 (Maersk Lanco)",
        "base_imponible": 34117.20, "pct_iva": 4, "cuota_iva": 1364.69,
        "total": 35481.89, "cobrada": "Si", "forma_pago": "Transferencia",
        "observaciones": "1er envio - 19,440 kg limon Genova"
    },
    {
        "fecha": "25/08/2025", "trimestre": "T3", "num_factura": "INV/2025/00002",
        "cliente": "MALAGA NATURAL 2012 S.L.", "cif_cliente": "B93159044",
        "concepto": "Venta limon Lisboa - Containers SUDU8020260 + MNBU0233449",
        "base_imponible": 34933.68, "pct_iva": 4, "cuota_iva": 1397.34,
        "total": 36331.02, "cobrada": "Si", "forma_pago": "Transferencia",
        "observaciones": "2do envio - 19,440 kg limon Lisboa"
    },
    {
        "fecha": "25/08/2025", "trimestre": "T3", "num_factura": "INV/2025/00003",
        "cliente": "MALAGA NATURAL 2012 S.L.", "cif_cliente": "B93159044",
        "concepto": "Venta limon - Containers MNBU0233449 + SUDU6177121",
        "base_imponible": 34992.00, "pct_iva": 4, "cuota_iva": 1399.69,
        "total": 36391.69, "cobrada": "Si", "forma_pago": "Transferencia",
        "observaciones": "3er envio - 19,440 kg limon"
    },
    {
        "fecha": "25/08/2025", "trimestre": "T3", "num_factura": "INV/2025/00004",
        "cliente": "MALAGA NATURAL 2012 S.L.", "cif_cliente": "B93159044",
        "concepto": "Venta limon - Container SUDU6177121 (Maersk Londrina)",
        "base_imponible": 34933.68, "pct_iva": 4, "cuota_iva": 1397.35,
        "total": 36331.03, "cobrada": "Si", "forma_pago": "Transferencia",
        "observaciones": "4to envio - 19,440 kg limon"
    },
]

# --- FACTURAS RECIBIDAS (Gastos / Proveedores) ---
facturas_recibidas = [
    # === CAUQUEN ARGENTINA - Compra de limones (extranjera, sin IVA espanol) ===
    {
        "fecha": "28/07/2025", "trimestre": "T3", "num_factura": "E 00005-00004696",
        "proveedor": "CAUQUEN ARGENTINA S.A.U.", "cif_proveedor": "AR 30-70793648-3",
        "concepto": "Compra 1440 cajas limon fresco FOB Argentina - MNBU4144463",
        "base_imponible": 28800.00, "moneda": "USD", "pct_iva": 0, "cuota_iva": 0,
        "total": 28800.00, "tipo_gasto": "Compra mercaderia (importacion)",
        "pagada": "Si", "forma_pago": "Transferencia internacional",
        "observaciones": "Vessel San Raphael Maersk. Pago via BECA PAS LLC, Bank of America"
    },

    # === LNET S.A. / LOGINET - Flete maritimo (extranjera Argentina, USD) ===
    {
        "fecha": "26/06/2025", "trimestre": "T2", "num_factura": "FC A 0003-00026826",
        "proveedor": "LNET S.A. (LOGINET)", "cif_proveedor": "AR 30-71035801-6",
        "concepto": "Flete maritimo SUDU8020260 - Maersk Lanco (Buenos Aires > Sines PT)",
        "base_imponible": 3900.00, "moneda": "USD", "pct_iva": 0, "cuota_iva": 0,
        "total": 3900.00, "tipo_gasto": "Transporte maritimo (importacion)",
        "pagada": "Si", "forma_pago": "Transferencia internacional",
        "observaciones": "ETD 31/05/25. Pago a Amerant Bank, Miami"
    },
    {
        "fecha": "26/06/2025", "trimestre": "T2", "num_factura": "NC B 0003-00003967",
        "proveedor": "LNET S.A. (LOGINET)", "cif_proveedor": "AR 30-71035801-6",
        "concepto": "Nota de credito - anula FC anterior SUDU8020260",
        "base_imponible": -3900.00, "moneda": "USD", "pct_iva": 0, "cuota_iva": 0,
        "total": -3900.00, "tipo_gasto": "Transporte maritimo (importacion)",
        "pagada": "Si", "forma_pago": "Compensacion",
        "observaciones": "Aplica sobre FC B 0003-00026535"
    },
    {
        "fecha": "27/06/2025", "trimestre": "T2", "num_factura": "FC B 0003-00026855",
        "proveedor": "LNET S.A. (LOGINET)", "cif_proveedor": "AR 30-71035801-6",
        "concepto": "Flete maritimo MNBU0233449 - San Lorenzo Maersk (Buenos Aires > Sines PT)",
        "base_imponible": 3900.00, "moneda": "USD", "pct_iva": 0, "cuota_iva": 0,
        "total": 3900.00, "tipo_gasto": "Transporte maritimo (importacion)",
        "pagada": "Si", "forma_pago": "Transferencia internacional",
        "observaciones": "ETD 07/06/25"
    },
    {
        "fecha": "07/07/2025", "trimestre": "T3", "num_factura": "FC B 0003-00027144",
        "proveedor": "LNET S.A. (LOGINET)", "cif_proveedor": "AR 30-71035801-6",
        "concepto": "Flete maritimo SUDU6177121 - Maersk Londrina (Buenos Aires > Sines PT)",
        "base_imponible": 3900.00, "moneda": "USD", "pct_iva": 0, "cuota_iva": 0,
        "total": 3900.00, "tipo_gasto": "Transporte maritimo (importacion)",
        "pagada": "Si", "forma_pago": "Transferencia internacional",
        "observaciones": "ETD 21/06/25"
    },
    {
        "fecha": "05/08/2025", "trimestre": "T3", "num_factura": "FC A 0003-00027760",
        "proveedor": "LNET S.A. (LOGINET)", "cif_proveedor": "AR 30-71035801-6",
        "concepto": "Flete maritimo MNBU4144463 - San Raphael Maersk (Buenos Aires > Lisboa PT)",
        "base_imponible": 3900.00, "moneda": "USD", "pct_iva": 0, "cuota_iva": 0,
        "total": 3900.00, "tipo_gasto": "Transporte maritimo (importacion)",
        "pagada": "Si", "forma_pago": "Transferencia internacional",
        "observaciones": "ETD 26/07/25"
    },

    # === PRIMAFRIO S.L. - Transporte frigorifico nacional ===
    {
        "fecha": "30/07/2025", "trimestre": "T3", "num_factura": "1090161585",
        "proveedor": "PRIMAFRIO S.L.", "cif_proveedor": "B73047599",
        "concepto": "Transporte frigorifico",
        "base_imponible": 800.00, "moneda": "EUR", "pct_iva": 21, "cuota_iva": 168.00,
        "total": 968.00, "tipo_gasto": "Transporte nacional",
        "pagada": "Pendiente", "forma_pago": "Transferencia 30 dias",
        "observaciones": "Vto. 29/08/2025"
    },
    {
        "fecha": "06/08/2025", "trimestre": "T3", "num_factura": "1090165048",
        "proveedor": "PRIMAFRIO S.L.", "cif_proveedor": "B73047599",
        "concepto": "Transporte frigorifico",
        "base_imponible": 800.00, "moneda": "EUR", "pct_iva": 21, "cuota_iva": 168.00,
        "total": 968.00, "tipo_gasto": "Transporte nacional",
        "pagada": "Pendiente", "forma_pago": "Transferencia 30 dias",
        "observaciones": "Vto. 05/09/2025"
    },
    {
        "fecha": "08/08/2025", "trimestre": "T3", "num_factura": "1090167087",
        "proveedor": "PRIMAFRIO S.L.", "cif_proveedor": "B73047599",
        "concepto": "Transporte frigorifico",
        "base_imponible": 800.00, "moneda": "EUR", "pct_iva": 21, "cuota_iva": 168.00,
        "total": 968.00, "tipo_gasto": "Transporte nacional",
        "pagada": "Pendiente", "forma_pago": "Transferencia 30 dias",
        "observaciones": "Vto. 07/09/2025"
    },
    {
        "fecha": "12/08/2025", "trimestre": "T3", "num_factura": "1090168047",
        "proveedor": "PRIMAFRIO S.L.", "cif_proveedor": "B73047599",
        "concepto": "Transporte frigorifico",
        "base_imponible": 800.00, "moneda": "EUR", "pct_iva": 21, "cuota_iva": 168.00,
        "total": 968.00, "tipo_gasto": "Transporte nacional",
        "pagada": "Pendiente", "forma_pago": "Transferencia 30 dias",
        "observaciones": "Vto. 11/09/2025"
    },

    # === PRIMATRANSIT S.L. - Despacho aduanero + gastos importacion ===
    {
        "fecha": "27/08/2025", "trimestre": "T3", "num_factura": "2390101398",
        "proveedor": "PRIMATRANSIT S.L.", "cif_proveedor": "B16815003",
        "concepto": "Despacho aduanero SUDU8020260 (aranceles, IVA aduana, fito, caucion)",
        "base_imponible": 7608.29, "moneda": "EUR", "pct_iva": 0, "cuota_iva": 0,
        "iva_mixto": "5,797.04 exento + 1,811.25 al 21% (380.36€ IVA)",
        "total": 7988.65, "tipo_gasto": "Despacho aduanero (importacion)",
        "pagada": "Pendiente", "forma_pago": "Transferencia 30 dias",
        "observaciones": "Incluye derechos arancel 2,426.94€, IVA aduana 2,444.89€, tasas fito 120€, caucion 50€, certificados 70€"
    },
    {
        "fecha": "27/08/2025", "trimestre": "T3", "num_factura": "2390101399",
        "proveedor": "PRIMATRANSIT S.L.", "cif_proveedor": "B16815003",
        "concepto": "Despacho aduanero MNBU0233449 (aranceles, IVA aduana, fito)",
        "base_imponible": 7608.29, "moneda": "EUR", "pct_iva": 0, "cuota_iva": 0,
        "iva_mixto": "5,797.04 exento + 1,811.25 al 21% (380.36€ IVA)",
        "total": 7988.65, "tipo_gasto": "Despacho aduanero (importacion)",
        "pagada": "Pendiente", "forma_pago": "Transferencia 30 dias",
        "observaciones": "Incluye derechos arancel 2,830.05 + 2,846.99€, tasas fito 120€"
    },
    {
        "fecha": "27/08/2025", "trimestre": "T3", "num_factura": "2390101400",
        "proveedor": "PRIMATRANSIT S.L.", "cif_proveedor": "B16815003",
        "concepto": "Despacho aduanero SUDU6177121 (aranceles, IVA aduana, fito)",
        "base_imponible": 7372.26, "moneda": "EUR", "pct_iva": 0, "cuota_iva": 0,
        "iva_mixto": "5,603.51 exento + 1,768.75 al 21% (371.44€ IVA)",
        "total": 7743.70, "tipo_gasto": "Despacho aduanero (importacion)",
        "pagada": "Pendiente", "forma_pago": "Transferencia 30 dias",
        "observaciones": "Contenedor SUDU6177121 - Maersk Londrina"
    },
    {
        "fecha": "17/09/2025", "trimestre": "T3", "num_factura": "2390101460",
        "proveedor": "PRIMATRANSIT S.L.", "cif_proveedor": "B16815003",
        "concepto": "Despacho + flete + costes MNBU4144463 (naviera, aduana, fito, transporte)",
        "base_imponible": 11040.63, "moneda": "EUR", "pct_iva": 0, "cuota_iva": 0,
        "iva_mixto": "9,848.33 exento + 1,192.30 al 21% (250.38€ IVA)",
        "total": 11291.01, "tipo_gasto": "Despacho aduanero (importacion)",
        "pagada": "Pendiente", "forma_pago": "Transferencia 30 dias",
        "observaciones": "Flete Maersk 2,345.01€, costes naviera 195€, derechos, IVA aduana, transporte"
    },
    {
        "fecha": "30/12/2025", "trimestre": "T4", "num_factura": "2390102446",
        "proveedor": "PRIMATRANSIT S.L.", "cif_proveedor": "B16815003",
        "concepto": "Gastos logisticos adicionales",
        "base_imponible": 1880.30, "moneda": "EUR", "pct_iva": 0, "cuota_iva": 0,
        "iva_mixto": "688.00 exento + 1,192.30 al 21%",
        "total": 1880.30, "tipo_gasto": "Despacho aduanero (importacion)",
        "pagada": "Pendiente", "forma_pago": "Transferencia 30 dias",
        "observaciones": "Vto. 29/01/2026"
    },
    {
        "fecha": "30/12/2025", "trimestre": "T4", "num_factura": "2390400070",
        "proveedor": "PRIMATRANSIT S.L.", "cif_proveedor": "B16815003",
        "concepto": "Abono suplidos MNBU4144463 (ref factura 2390101702)",
        "base_imponible": -453.00, "moneda": "EUR", "pct_iva": 0, "cuota_iva": 0,
        "total": -453.00, "tipo_gasto": "Despacho aduanero (importacion)",
        "pagada": "Pendiente", "forma_pago": "Compensacion",
        "observaciones": "Abono sobre factura 2390101702"
    },

    # === TRANSITAINER PORTUGAL LTDA - Despacho aduanero en Lisboa ===
    {
        "fecha": "14/08/2025", "trimestre": "T3", "num_factura": "TT FT25S/000116",
        "proveedor": "TRANSITAINER PORTUGAL LTDA", "cif_proveedor": "PT (extranjera)",
        "concepto": "Despacho aduanero MNBU0233449 en Lisboa (aranceles, IVA import, fito, transporte)",
        "base_imponible": 6908.82, "moneda": "EUR", "pct_iva": 0, "cuota_iva": 0,
        "total": 6908.82, "tipo_gasto": "Despacho aduanero (importacion)",
        "pagada": "Pendiente", "forma_pago": "Transferencia",
        "observaciones": "Arancel 2,830.05€, IVA import 2,846.99€, caucion 56.78€, fito 100€, certif 50€, despacho 150€, transp PIFF 100€, PIFF descons 660€, guia emolumentos 65€, gestion 50€"
    },

    # === MAERSK - Naviera (facturas con IVA 0% por ser operador extranjero) ===
    {
        "fecha": "06/07/2025", "trimestre": "T3", "num_factura": "7532745537",
        "proveedor": "MAERSK SPAIN SLU (agente Maersk A/S)", "cif_proveedor": "DK53139655",
        "concepto": "Importacion contenedor SUDU8020260 - Green Pole 526N (Buenos Aires > Sines)",
        "base_imponible": 0, "moneda": "EUR", "pct_iva": 0, "cuota_iva": 0,
        "total": 0, "tipo_gasto": "Gastos naviera (importacion)",
        "pagada": "Pendiente", "forma_pago": "Transferencia",
        "observaciones": "Importe pendiente verificar (PDF parcialmente legible)"
    },
    {
        "fecha": "18/07/2025", "trimestre": "T3", "num_factura": "7533441178",
        "proveedor": "MAERSK SPAIN SLU", "cif_proveedor": "DK53139655",
        "concepto": "Ocupaciones contenedor SUDU8020260 (Green Pole 526N)",
        "base_imponible": 0, "moneda": "EUR", "pct_iva": 0, "cuota_iva": 0,
        "total": 0, "tipo_gasto": "Gastos naviera (importacion)",
        "pagada": "Pendiente", "forma_pago": "Transferencia",
        "observaciones": "Ocupacion en Sines/Lisboa"
    },
    {
        "fecha": "18/07/2025", "trimestre": "T3", "num_factura": "7533441271",
        "proveedor": "MAERSK SPAIN SLU", "cif_proveedor": "DK53139655",
        "concepto": "Demoras y paralizaciones SUDU8020260 (Green Pole 526N)",
        "base_imponible": 0, "moneda": "EUR", "pct_iva": 0, "cuota_iva": 0,
        "total": 0, "tipo_gasto": "Gastos naviera (importacion)",
        "pagada": "Pendiente", "forma_pago": "Transferencia",
        "observaciones": "Demoras contenedor en Sines/Lisboa"
    },
    {
        "fecha": "18/07/2025", "trimestre": "T3", "num_factura": "7533442509",
        "proveedor": "MAERSK SPAIN SLU", "cif_proveedor": "DK53139655",
        "concepto": "Importacion contenedor MNBU0233449 - Green Pole 526N",
        "base_imponible": 0, "moneda": "EUR", "pct_iva": 0, "cuota_iva": 0,
        "total": 0, "tipo_gasto": "Gastos naviera (importacion)",
        "pagada": "Pendiente", "forma_pago": "Transferencia",
        "observaciones": "Buenos Aires > Sines"
    },
    {
        "fecha": "18/07/2025", "trimestre": "T3", "num_factura": "7533463809",
        "proveedor": "MAERSK SPAIN SLU", "cif_proveedor": "DK53139655",
        "concepto": "Ocupaciones contenedor MNBU0233449 (Green Pole 527N)",
        "base_imponible": 0, "moneda": "EUR", "pct_iva": 0, "cuota_iva": 0,
        "total": 0, "tipo_gasto": "Gastos naviera (importacion)",
        "pagada": "Pendiente", "forma_pago": "Transferencia",
        "observaciones": ""
    },
    {
        "fecha": "18/07/2025", "trimestre": "T3", "num_factura": "7533463896",
        "proveedor": "MAERSK SPAIN SLU", "cif_proveedor": "DK53139655",
        "concepto": "Demoras y paralizaciones MNBU0233449",
        "base_imponible": 0, "moneda": "EUR", "pct_iva": 0, "cuota_iva": 0,
        "total": 0, "tipo_gasto": "Gastos naviera (importacion)",
        "pagada": "Pendiente", "forma_pago": "Transferencia",
        "observaciones": ""
    },
    {
        "fecha": "22/09/2025", "trimestre": "T3", "num_factura": "7537075071",
        "proveedor": "MAERSK A/S (via Maersk Benelux Rotterdam)", "cif_proveedor": "DK53139655",
        "concepto": "Importacion MNBU4144463 - Maersk Lanco 531N (Buenos Aires > Rotterdam)",
        "base_imponible": 453.00, "moneda": "EUR", "pct_iva": 0, "cuota_iva": 0,
        "total": 453.00, "tipo_gasto": "Gastos naviera (importacion)",
        "pagada": "Pendiente", "forma_pago": "Transferencia",
        "observaciones": "Container Protect 28€ + Documentation 50€ + THC Destino 375€. Factura a CSI-FRESH BV (Holanda)"
    },

    # === ODOO - Software ERP ===
    {
        "fecha": "01/07/2025", "trimestre": "T3", "num_factura": "2025/07/000004",
        "proveedor": "ODOO S.A.", "cif_proveedor": "BE0477472701",
        "concepto": "Suscripcion Odoo Enterprise (julio 2025)",
        "base_imponible": 14.80, "moneda": "EUR", "pct_iva": 0, "cuota_iva": 0,
        "total": 14.80, "tipo_gasto": "Software / servicios informaticos",
        "pagada": "Si", "forma_pago": "Tarjeta",
        "observaciones": "Intracomunitaria - autoliquidacion IVA. Belgica VAT BE0477472701"
    },
    {
        "fecha": "01/09/2025", "trimestre": "T3", "num_factura": "2025/09/000020",
        "proveedor": "ODOO S.A.", "cif_proveedor": "BE0477472701",
        "concepto": "Suscripcion Odoo Enterprise (septiembre 2025)",
        "base_imponible": 14.80, "moneda": "EUR", "pct_iva": 0, "cuota_iva": 0,
        "total": 14.80, "tipo_gasto": "Software / servicios informaticos",
        "pagada": "Si", "forma_pago": "Tarjeta",
        "observaciones": "Intracomunitaria - autoliquidacion IVA. Belgica"
    },
    {
        "fecha": "09/10/2025", "trimestre": "T4", "num_factura": "INV/2025/10/00669",
        "proveedor": "ODOO ERP SP SL", "cif_proveedor": "B72659014",
        "concepto": "Suscripcion Odoo Enterprise (octubre 2025)",
        "base_imponible": 14.80, "moneda": "EUR", "pct_iva": 21, "cuota_iva": 3.11,
        "total": 17.91, "tipo_gasto": "Software / servicios informaticos",
        "pagada": "Si", "forma_pago": "Tarjeta",
        "observaciones": "Facturado desde Odoo Espana (Valencia)"
    },

    # === SOCIEDAD ANDALUZA DE PUBLICIDAD (COPYRAP) - Tarjetas ===
    {
        "fecha": "30/09/2025", "trimestre": "T3", "num_factura": "T-AUR-25/023355",
        "proveedor": "SOCIEDAD ANDALUZA DE PUBLICIDAD Y MARKETING SL", "cif_proveedor": "B02875516",
        "concepto": "100 tarjetas de visita 1+1 papel estucado mate (x2)",
        "base_imponible": 35.06, "moneda": "EUR", "pct_iva": 21, "cuota_iva": 7.36,
        "total": 42.42, "tipo_gasto": "Publicidad y marketing",
        "pagada": "Si", "forma_pago": "Tarjeta",
        "observaciones": "Copyrap Malaga"
    },

    # === EL CORTE INGLES - Material oficina ===
    {
        "fecha": "10/09/2025", "trimestre": "T3", "num_factura": "06975016374",
        "proveedor": "EL CORTE INGLES S.A.", "cif_proveedor": "A28017895",
        "concepto": "Material oficina (tope puerta, otros)",
        "base_imponible": 13.21, "moneda": "EUR", "pct_iva": 21, "cuota_iva": 2.78,
        "total": 15.99, "tipo_gasto": "Material oficina",
        "pagada": "Si", "forma_pago": "Tarjeta",
        "observaciones": "Tienda Marbella"
    },
    {
        "fecha": "15/09/2025", "trimestre": "T3", "num_factura": "06975016660",
        "proveedor": "EL CORTE INGLES S.A.", "cif_proveedor": "A28017895",
        "concepto": "Material oficina (bolsas, tealights, otros)",
        "base_imponible": 0.17, "moneda": "EUR", "pct_iva": 21, "cuota_iva": 0.03,
        "total": 0.20, "tipo_gasto": "Material oficina",
        "pagada": "Si", "forma_pago": "Tarjeta",
        "observaciones": "Tienda Marbella - importe parcial visible"
    },
]

# ============================================================
# GENERAR EXCEL
# ============================================================

def crear_estilo_header():
    return {
        "font": Font(bold=True, color="FFFFFF", size=10),
        "fill": PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
    }

def crear_estilo_datos():
    return {
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        ),
        "alignment": Alignment(vertical="center", wrap_text=True)
    }

def aplicar_estilos_header(ws, fila, num_cols):
    estilos = crear_estilo_header()
    for col in range(1, num_cols + 1):
        celda = ws.cell(row=fila, column=col)
        celda.font = estilos["font"]
        celda.fill = estilos["fill"]
        celda.alignment = estilos["alignment"]
        celda.border = estilos["border"]

def aplicar_estilos_datos(ws, fila, num_cols):
    estilos = crear_estilo_datos()
    for col in range(1, num_cols + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = estilos["border"]
        celda.alignment = estilos["alignment"]

def formato_moneda(ws, fila, columnas):
    for col in columnas:
        celda = ws.cell(row=fila, column=col)
        celda.number_format = '#,##0.00 €'

def auto_ancho(ws):
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 45)
        ws.column_dimensions[column_letter].width = max(adjusted_width, 12)


def main():
    wb = openpyxl.Workbook()

    # ===== PESTANA 1: INGRESOS (Facturas emitidas) =====
    ws_ingresos = wb.active
    ws_ingresos.title = "Ingresos"
    headers_ingresos = [
        "Fecha", "Trimestre", "N Factura", "Cliente", "CIF Cliente",
        "Concepto", "Base Imponible", "% IVA", "Cuota IVA",
        "Total Factura", "Cobrada", "Forma Pago", "Observaciones"
    ]
    for col, h in enumerate(headers_ingresos, 1):
        ws_ingresos.cell(row=1, column=col, value=h)
    aplicar_estilos_header(ws_ingresos, 1, len(headers_ingresos))

    for i, f in enumerate(facturas_emitidas, 2):
        ws_ingresos.cell(row=i, column=1, value=f["fecha"])
        ws_ingresos.cell(row=i, column=2, value=f["trimestre"])
        ws_ingresos.cell(row=i, column=3, value=f["num_factura"])
        ws_ingresos.cell(row=i, column=4, value=f["cliente"])
        ws_ingresos.cell(row=i, column=5, value=f["cif_cliente"])
        ws_ingresos.cell(row=i, column=6, value=f["concepto"])
        ws_ingresos.cell(row=i, column=7, value=f["base_imponible"])
        ws_ingresos.cell(row=i, column=8, value=f["pct_iva"])
        ws_ingresos.cell(row=i, column=9, value=f["cuota_iva"])
        ws_ingresos.cell(row=i, column=10, value=f["total"])
        ws_ingresos.cell(row=i, column=11, value=f["cobrada"])
        ws_ingresos.cell(row=i, column=12, value=f["forma_pago"])
        ws_ingresos.cell(row=i, column=13, value=f["observaciones"])
        aplicar_estilos_datos(ws_ingresos, i, len(headers_ingresos))
        formato_moneda(ws_ingresos, i, [7, 9, 10])

    # Fila totales
    fila_total = len(facturas_emitidas) + 2
    ws_ingresos.cell(row=fila_total, column=6, value="TOTALES").font = Font(bold=True)
    ws_ingresos.cell(row=fila_total, column=7, value=sum(f["base_imponible"] for f in facturas_emitidas))
    ws_ingresos.cell(row=fila_total, column=9, value=sum(f["cuota_iva"] for f in facturas_emitidas))
    ws_ingresos.cell(row=fila_total, column=10, value=sum(f["total"] for f in facturas_emitidas))
    for col in [7, 9, 10]:
        ws_ingresos.cell(row=fila_total, column=col).font = Font(bold=True)
        ws_ingresos.cell(row=fila_total, column=col).number_format = '#,##0.00 €'
    auto_ancho(ws_ingresos)

    # ===== PESTANA 2: GASTOS (Facturas recibidas) =====
    ws_gastos = wb.create_sheet("Gastos")
    headers_gastos = [
        "Fecha", "Trimestre", "N Factura", "Proveedor", "CIF Proveedor",
        "Concepto", "Base Imponible", "Moneda", "% IVA", "Cuota IVA",
        "Total Factura", "Tipo Gasto", "Pagada", "Forma Pago", "Observaciones"
    ]
    for col, h in enumerate(headers_gastos, 1):
        ws_gastos.cell(row=1, column=col, value=h)
    aplicar_estilos_header(ws_gastos, 1, len(headers_gastos))

    for i, f in enumerate(facturas_recibidas, 2):
        ws_gastos.cell(row=i, column=1, value=f["fecha"])
        ws_gastos.cell(row=i, column=2, value=f["trimestre"])
        ws_gastos.cell(row=i, column=3, value=f["num_factura"])
        ws_gastos.cell(row=i, column=4, value=f["proveedor"])
        ws_gastos.cell(row=i, column=5, value=f["cif_proveedor"])
        ws_gastos.cell(row=i, column=6, value=f["concepto"])
        ws_gastos.cell(row=i, column=7, value=f["base_imponible"])
        ws_gastos.cell(row=i, column=8, value=f.get("moneda", "EUR"))
        ws_gastos.cell(row=i, column=9, value=f["pct_iva"])
        ws_gastos.cell(row=i, column=10, value=f["cuota_iva"])
        ws_gastos.cell(row=i, column=11, value=f["total"])
        ws_gastos.cell(row=i, column=12, value=f["tipo_gasto"])
        ws_gastos.cell(row=i, column=13, value=f.get("pagada", ""))
        ws_gastos.cell(row=i, column=14, value=f.get("forma_pago", ""))
        ws_gastos.cell(row=i, column=15, value=f["observaciones"])
        aplicar_estilos_datos(ws_gastos, i, len(headers_gastos))
        formato_moneda(ws_gastos, i, [7, 10, 11])

    fila_total = len(facturas_recibidas) + 2
    ws_gastos.cell(row=fila_total, column=6, value="TOTALES (EUR)").font = Font(bold=True)
    gastos_eur = [f for f in facturas_recibidas if f.get("moneda") == "EUR"]
    ws_gastos.cell(row=fila_total, column=7, value=sum(f["base_imponible"] for f in gastos_eur))
    ws_gastos.cell(row=fila_total, column=10, value=sum(f["cuota_iva"] for f in gastos_eur))
    ws_gastos.cell(row=fila_total, column=11, value=sum(f["total"] for f in gastos_eur))
    for col in [7, 10, 11]:
        ws_gastos.cell(row=fila_total, column=col).font = Font(bold=True)
        ws_gastos.cell(row=fila_total, column=col).number_format = '#,##0.00 €'

    fila_total2 = fila_total + 1
    ws_gastos.cell(row=fila_total2, column=6, value="TOTALES (USD)").font = Font(bold=True)
    gastos_usd = [f for f in facturas_recibidas if f.get("moneda") == "USD"]
    ws_gastos.cell(row=fila_total2, column=7, value=sum(f["base_imponible"] for f in gastos_usd))
    ws_gastos.cell(row=fila_total2, column=11, value=sum(f["total"] for f in gastos_usd))
    for col in [7, 11]:
        ws_gastos.cell(row=fila_total2, column=col).font = Font(bold=True)
        ws_gastos.cell(row=fila_total2, column=col).number_format = '#,##0.00 $'
    auto_ancho(ws_gastos)

    # ===== PESTANA 3: REGISTRO FACT. EMITIDAS (formato fiscal) =====
    ws_reg_emit = wb.create_sheet("Registro Fact. Emitidas")
    headers_reg_emit = [
        "Fecha Expedicion", "Fecha Operacion", "Trimestre", "N Factura",
        "Cliente", "CIF Cliente", "Base Imponible Exenta",
        "Base Imponible 4%", "Cuota IVA 4%",
        "Base Imponible 10%", "Cuota IVA 10%",
        "Base Imponible 21%", "Cuota IVA 21%",
        "Total Factura", "Observaciones"
    ]
    for col, h in enumerate(headers_reg_emit, 1):
        ws_reg_emit.cell(row=1, column=col, value=h)
    aplicar_estilos_header(ws_reg_emit, 1, len(headers_reg_emit))

    for i, f in enumerate(facturas_emitidas, 2):
        ws_reg_emit.cell(row=i, column=1, value=f["fecha"])
        ws_reg_emit.cell(row=i, column=2, value=f["fecha"])
        ws_reg_emit.cell(row=i, column=3, value=f["trimestre"])
        ws_reg_emit.cell(row=i, column=4, value=f["num_factura"])
        ws_reg_emit.cell(row=i, column=5, value=f["cliente"])
        ws_reg_emit.cell(row=i, column=6, value=f["cif_cliente"])
        ws_reg_emit.cell(row=i, column=7, value=0)
        ws_reg_emit.cell(row=i, column=8, value=f["base_imponible"])
        ws_reg_emit.cell(row=i, column=9, value=f["cuota_iva"])
        ws_reg_emit.cell(row=i, column=10, value=0)
        ws_reg_emit.cell(row=i, column=11, value=0)
        ws_reg_emit.cell(row=i, column=12, value=0)
        ws_reg_emit.cell(row=i, column=13, value=0)
        ws_reg_emit.cell(row=i, column=14, value=f["total"])
        ws_reg_emit.cell(row=i, column=15, value=f["observaciones"])
        aplicar_estilos_datos(ws_reg_emit, i, len(headers_reg_emit))
        formato_moneda(ws_reg_emit, i, [7, 8, 9, 10, 11, 12, 13, 14])
    auto_ancho(ws_reg_emit)

    # ===== PESTANA 4: REGISTRO FACT. RECIBIDAS (formato fiscal) =====
    ws_reg_rec = wb.create_sheet("Registro Fact. Recibidas")
    headers_reg_rec = [
        "Fecha Recepcion", "Fecha Factura", "Trimestre", "N Factura",
        "Proveedor", "CIF Proveedor", "Moneda",
        "Base Imponible 0%/Exenta", "Base Imponible 4%", "Cuota IVA 4%",
        "Base Imponible 21%", "Cuota IVA 21%",
        "IVA Deducible", "Total Factura",
        "Tipo Gasto", "Observaciones"
    ]
    for col, h in enumerate(headers_reg_rec, 1):
        ws_reg_rec.cell(row=1, column=col, value=h)
    aplicar_estilos_header(ws_reg_rec, 1, len(headers_reg_rec))

    for i, f in enumerate(facturas_recibidas, 2):
        ws_reg_rec.cell(row=i, column=1, value=f["fecha"])
        ws_reg_rec.cell(row=i, column=2, value=f["fecha"])
        ws_reg_rec.cell(row=i, column=3, value=f["trimestre"])
        ws_reg_rec.cell(row=i, column=4, value=f["num_factura"])
        ws_reg_rec.cell(row=i, column=5, value=f["proveedor"])
        ws_reg_rec.cell(row=i, column=6, value=f["cif_proveedor"])
        ws_reg_rec.cell(row=i, column=7, value=f.get("moneda", "EUR"))

        # Clasificar por tipo de IVA
        if f["pct_iva"] == 0:
            ws_reg_rec.cell(row=i, column=8, value=f["base_imponible"])  # Exenta/0%
            ws_reg_rec.cell(row=i, column=9, value=0)
            ws_reg_rec.cell(row=i, column=10, value=0)
            ws_reg_rec.cell(row=i, column=11, value=0)
            ws_reg_rec.cell(row=i, column=12, value=0)
        elif f["pct_iva"] == 4:
            ws_reg_rec.cell(row=i, column=8, value=0)
            ws_reg_rec.cell(row=i, column=9, value=f["base_imponible"])
            ws_reg_rec.cell(row=i, column=10, value=f["cuota_iva"])
            ws_reg_rec.cell(row=i, column=11, value=0)
            ws_reg_rec.cell(row=i, column=12, value=0)
        elif f["pct_iva"] == 21:
            ws_reg_rec.cell(row=i, column=8, value=0)
            ws_reg_rec.cell(row=i, column=9, value=0)
            ws_reg_rec.cell(row=i, column=10, value=0)
            ws_reg_rec.cell(row=i, column=11, value=f["base_imponible"])
            ws_reg_rec.cell(row=i, column=12, value=f["cuota_iva"])

        ws_reg_rec.cell(row=i, column=13, value=f["cuota_iva"])  # IVA deducible
        ws_reg_rec.cell(row=i, column=14, value=f["total"])
        ws_reg_rec.cell(row=i, column=15, value=f["tipo_gasto"])
        ws_reg_rec.cell(row=i, column=16, value=f["observaciones"])
        aplicar_estilos_datos(ws_reg_rec, i, len(headers_reg_rec))
        formato_moneda(ws_reg_rec, i, [8, 9, 10, 11, 12, 13, 14])
    auto_ancho(ws_reg_rec)

    # ===== PESTANA 5: RESUMEN TRIMESTRAL =====
    ws_resumen = wb.create_sheet("Resumen Trimestral")
    headers_resumen = [
        "Trimestre", "Ingresos Base", "IVA Repercutido (4%)",
        "Gastos Base (EUR)", "Gastos Base (USD)", "IVA Soportado Deducible",
        "Resultado IVA (303)", "Observaciones"
    ]
    for col, h in enumerate(headers_resumen, 1):
        ws_resumen.cell(row=1, column=col, value=h)
    aplicar_estilos_header(ws_resumen, 1, len(headers_resumen))

    for idx, trim in enumerate(["T1", "T2", "T3", "T4"], 2):
        ingresos_trim = [f for f in facturas_emitidas if f["trimestre"] == trim]
        gastos_eur_trim = [f for f in facturas_recibidas if f["trimestre"] == trim and f.get("moneda") == "EUR"]
        gastos_usd_trim = [f for f in facturas_recibidas if f["trimestre"] == trim and f.get("moneda") == "USD"]

        base_ingresos = sum(f["base_imponible"] for f in ingresos_trim)
        iva_repercutido = sum(f["cuota_iva"] for f in ingresos_trim)
        base_gastos_eur = sum(f["base_imponible"] for f in gastos_eur_trim)
        base_gastos_usd = sum(f["base_imponible"] for f in gastos_usd_trim)
        iva_soportado = sum(f["cuota_iva"] for f in gastos_eur_trim)
        resultado_iva = iva_repercutido - iva_soportado

        ws_resumen.cell(row=idx, column=1, value=trim)
        ws_resumen.cell(row=idx, column=2, value=base_ingresos)
        ws_resumen.cell(row=idx, column=3, value=iva_repercutido)
        ws_resumen.cell(row=idx, column=4, value=base_gastos_eur)
        ws_resumen.cell(row=idx, column=5, value=base_gastos_usd)
        ws_resumen.cell(row=idx, column=6, value=iva_soportado)
        ws_resumen.cell(row=idx, column=7, value=resultado_iva)
        aplicar_estilos_datos(ws_resumen, idx, len(headers_resumen))
        formato_moneda(ws_resumen, idx, [2, 3, 4, 6, 7])

    # Totales
    fila_total = 6
    ws_resumen.cell(row=fila_total, column=1, value="TOTAL ANUAL").font = Font(bold=True, size=11)
    for col in [2, 3, 4, 5, 6, 7]:
        total = sum(
            ws_resumen.cell(row=r, column=col).value or 0
            for r in range(2, 6)
        )
        ws_resumen.cell(row=fila_total, column=col, value=total)
        ws_resumen.cell(row=fila_total, column=col).font = Font(bold=True)
        ws_resumen.cell(row=fila_total, column=col).number_format = '#,##0.00 €'
    auto_ancho(ws_resumen)

    # ===== PESTANA 6: PROVEEDORES =====
    ws_prov = wb.create_sheet("Proveedores")
    headers_prov = [
        "Proveedor", "CIF/NIF", "Pais", "Tipo", "Actividad",
        "Direccion", "Contacto", "Observaciones"
    ]
    for col, h in enumerate(headers_prov, 1):
        ws_prov.cell(row=1, column=col, value=h)
    aplicar_estilos_header(ws_prov, 1, len(headers_prov))

    proveedores = [
        ["CAUQUEN ARGENTINA S.A.U.", "30-70793648-3", "Argentina", "Exportador",
         "Produccion y exportacion de citricos", "La Pampa 1512, Buenos Aires",
         "+54 11 4787 3330", "Exportador principal de limones. Pago FOB en USD"],
        ["LNET S.A. (LOGINET)", "30-71035801-6", "Argentina", "Agente maritimo",
         "Logistica y transporte maritimo", "Paraguay 1225 piso 3, Buenos Aires",
         "info@loginetsa.com", "Flete maritimo Buenos Aires > Portugal. Pago USD via Amerant Bank Miami"],
        ["PRIMAFRIO S.L.", "B73047599", "Espana", "Transporte nacional",
         "Transporte frigorifico por carretera", "Autovia Mediterraneo km 596, Alhama de Murcia",
         "+34 968309187", "Transporte frigorifico Portugal > Mercamalaga"],
        ["PRIMATRANSIT S.L.", "B16815003", "Espana", "Agente aduanas",
         "Despacho aduanero y logistica internacional", "C/ Luxemburgo 3, Coslada, Madrid",
         "+34 968309187", "Agente de aduanas en Espana. Despachos, aranceles, IVA importacion"],
        ["TRANSITAINER PORTUGAL LTDA", "PT (pendiente)", "Portugal", "Agente aduanas",
         "Despacho aduanero en Portugal", "Portugal",
         "", "Despacho en puerto Lisboa/Sines"],
        ["MAERSK A/S / MAERSK SPAIN SLU", "DK53139655", "Dinamarca/Espana", "Naviera",
         "Transporte maritimo de contenedores", "Esplanaden 50, Copenhagen / P.E. La Finca, Pozuelo",
         "", "Naviera principal. Rutas Buenos Aires > Sines/Lisboa/Rotterdam"],
        ["ODOO S.A. / ODOO ERP SP SL", "BE0477472701 / B72659014", "Belgica/Espana", "Software",
         "ERP empresarial", "Chaussee de Namur 40, Belgium / Valencia, Espana",
         "info@odoo.com", "Suscripcion Odoo Enterprise mensual"],
        ["SOC. ANDALUZA DE PUBLICIDAD (COPYRAP)", "B02875516", "Espana", "Publicidad",
         "Impresion y marketing", "Avda. Aurora 12, Malaga",
         "951380761", "Tarjetas de visita"],
        ["EL CORTE INGLES S.A.", "A28017895", "Espana", "Comercio",
         "Gran almacen", "Marbella",
         "", "Compras menores material oficina"],
        ["COFACE", "W0012052G", "Francia/Espana", "Seguro credito",
         "Seguro de credito a la exportacion", "Via de los Poblados 3, Madrid",
         "91 310 42 24", "Seguro de credito sobre operaciones comerciales"],
    ]

    for i, p in enumerate(proveedores, 2):
        for col, val in enumerate(p, 1):
            ws_prov.cell(row=i, column=col, value=val)
        aplicar_estilos_datos(ws_prov, i, len(headers_prov))
    auto_ancho(ws_prov)

    # ===== PESTANA 7: CLIENTES =====
    ws_cli = wb.create_sheet("Clientes")
    headers_cli = ["Cliente", "CIF/NIF", "Direccion", "Actividad", "Observaciones"]
    for col, h in enumerate(headers_cli, 1):
        ws_cli.cell(row=1, column=col, value=h)
    aplicar_estilos_header(ws_cli, 1, len(headers_cli))

    clientes = [
        ["MALAGA NATURAL 2012 S.L.", "B93159044",
         "Mercamalaga, Avda. Ortega y Gasset 553, Modulos 142 y 806, 29196 Malaga",
         "Distribucion de frutas y verduras",
         "Cliente principal. Compra toda la produccion de limon importado. IVA 4%. Pago transferencia"],
    ]
    for i, c in enumerate(clientes, 2):
        for col, val in enumerate(c, 1):
            ws_cli.cell(row=i, column=col, value=val)
        aplicar_estilos_datos(ws_cli, i, len(headers_cli))
    auto_ancho(ws_cli)

    # ===== PESTANA 8: OPERACIONES IMPORTACION =====
    ws_import = wb.create_sheet("Operaciones Importacion")
    headers_import = [
        "Contenedor", "Buque", "ETD", "ETA", "Puerto Carga",
        "Puerto Descarga", "Bultos", "Kg Bruto", "Mercancia",
        "Exportador", "Flete USD", "DUA/DAU", "Observaciones"
    ]
    for col, h in enumerate(headers_import, 1):
        ws_import.cell(row=1, column=col, value=h)
    aplicar_estilos_header(ws_import, 1, len(headers_import))

    operaciones = [
        ["SUDU8020260", "MAERSK LANCO 526N", "31/05/2025", "01/07/2025",
         "Buenos Aires, AR", "Sines, PT", 1440, 28944, "Limon fresco",
         "CAUQUEN ARGENTINA", 3900, "DAU 40672_015272", "1er envio"],
        ["MNBU0233449", "SAN LORENZO MAERSK 523N", "07/06/2025", "09/07/2025",
         "Buenos Aires, AR", "Sines, PT", 1440, 28944, "Limon fresco",
         "CAUQUEN ARGENTINA", 3900, "PROB 40674_015150", "2do envio"],
        ["SUDU6177121", "MAERSK LONDRINA", "21/06/2025", "~17/07/2025",
         "Buenos Aires, AR", "Sines, PT", 1440, 28944, "Limon fresco",
         "CAUQUEN ARGENTINA", 3900, "DAU 40674_015149", "3er envio"],
        ["MNBU4144463", "SAN RAPHAEL MAERSK", "26/07/2025", "~01/09/2025",
         "Buenos Aires, AR", "Lisboa/Rotterdam", 1440, 28944, "Limon fresco",
         "CAUQUEN ARGENTINA", 3900, "Pendiente", "4to envio - factura Cauquen E 00005-00004696"],
    ]

    for i, op in enumerate(operaciones, 2):
        for col, val in enumerate(op, 1):
            ws_import.cell(row=i, column=col, value=val)
        aplicar_estilos_datos(ws_import, i, len(headers_import))
    auto_ancho(ws_import)

    # ===== PESTANA 9: CONCILIACION BANCARIA =====
    ws_banco = wb.create_sheet("Conciliacion Bancaria")
    headers_banco = [
        "Fecha Movimiento", "Trimestre", "Concepto Banco", "Importe",
        "Tipo", "Factura Asociada", "Proveedor/Cliente", "Conciliado", "Observaciones"
    ]
    for col, h in enumerate(headers_banco, 1):
        ws_banco.cell(row=1, column=col, value=h)
    aplicar_estilos_header(ws_banco, 1, len(headers_banco))

    movimientos = [
        ["09/07/2025", "T3", "Transferencia USD a LOGINET (flete maritimo)",
         -3900.00, "Gasto (USD)", "FC A 0003-00027760", "LNET S.A.",
         "Si", "Comision emision 31.63 USD + SWIFT 17.57 USD. Cargo CaixaBank"],
        ["11/07/2025", "T3", "Transferencia recibida Malaga Natural - proforma S00002",
         17971.20, "Ingreso", "Proforma S00002", "MALAGA NATURAL 2012 S.L.",
         "Si", "Anticipo compra limones 1440 cajas x 12€ + IVA 4%"],
    ]
    for i, m in enumerate(movimientos, 2):
        for col, val in enumerate(m, 1):
            ws_banco.cell(row=i, column=col, value=val)
        aplicar_estilos_datos(ws_banco, i, len(headers_banco))
    auto_ancho(ws_banco)

    # Guardar
    ruta_excel = "C:/Users/carli/PROYECTOS/CONTABILIDAD/clientes/pastorino-costa-del-sol/2025/libros_contables_2025.xlsx"
    wb.save(ruta_excel)
    print(f"Excel guardado en: {ruta_excel}")
    print(f"Pestanas: {wb.sheetnames}")
    print(f"Facturas emitidas: {len(facturas_emitidas)}")
    print(f"Facturas recibidas: {len(facturas_recibidas)}")


if __name__ == "__main__":
    main()
