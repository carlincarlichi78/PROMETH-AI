"""Tests para sfce/core/exportador.py — Exportador universal."""

import pytest
from pathlib import Path

from sfce.core.exportador import Exportador


@pytest.fixture
def exportador():
    return Exportador()


@pytest.fixture
def asientos_ejemplo():
    return [
        {
            "numero": 1, "fecha": "2025-01-15", "concepto": "Factura compra",
            "partidas": [
                {"subcuenta": "6000000001", "debe": 1000.0, "haber": 0, "concepto": "Gasto"},
                {"subcuenta": "4720000000", "debe": 210.0, "haber": 0, "concepto": "IVA"},
                {"subcuenta": "4000000001", "debe": 0, "haber": 1210.0, "concepto": "Proveedor"},
            ]
        },
        {
            "numero": 2, "fecha": "2025-01-20", "concepto": "Factura venta",
            "partidas": [
                {"subcuenta": "4300000001", "debe": 2420.0, "haber": 0, "concepto": "Cliente"},
                {"subcuenta": "7000000000", "debe": 0, "haber": 2000.0, "concepto": "Venta"},
                {"subcuenta": "4770000000", "debe": 0, "haber": 420.0, "concepto": "IVA rep"},
            ]
        },
    ]


class TestExportarCSV:
    def test_libro_diario_csv(self, exportador, asientos_ejemplo, tmp_path):
        ruta = tmp_path / "diario.csv"
        resultado = exportador.exportar_libro_diario_csv(asientos_ejemplo, ruta)
        assert resultado.exists()
        contenido = ruta.read_text(encoding="utf-8-sig")
        lineas = contenido.strip().split("\n")
        assert len(lineas) == 7  # 1 cabecera + 6 partidas
        assert "6000000001" in lineas[1]

    def test_facturas_recibidas_csv(self, exportador, tmp_path):
        facturas = [
            {"numero": "F-001", "fecha": "2025-01-15", "cif": "A11111111",
             "nombre": "Proveedor 1", "base": 1000, "iva": 210,
             "irpf": 0, "total": 1210, "pagada": True},
        ]
        ruta = tmp_path / "facturas.csv"
        resultado = exportador.exportar_facturas_csv(facturas, ruta, "recibidas")
        assert resultado.exists()
        contenido = ruta.read_text(encoding="utf-8-sig")
        assert "Proveedor 1" in contenido
        assert "Si" in contenido

    def test_facturas_emitidas_csv(self, exportador, tmp_path):
        facturas = [
            {"numero": "V-001", "fecha": "2025-02-01", "cif": "B22222222",
             "nombre": "Cliente 1", "base": 500, "iva": 105,
             "irpf": 0, "total": 605, "pagada": False},
        ]
        ruta = tmp_path / "emitidas.csv"
        resultado = exportador.exportar_facturas_csv(facturas, ruta, "emitidas")
        contenido = ruta.read_text(encoding="utf-8-sig")
        assert "CIF Receptor" in contenido
        assert "No" in contenido


class TestExportarExcel:
    def test_excel_multihoja(self, exportador, tmp_path):
        datos = {
            "Ventas": [
                {"Numero": "V-001", "Total": 605.0},
                {"Numero": "V-002", "Total": 1210.0},
            ],
            "Compras": [
                {"Numero": "F-001", "Total": 1210.0},
            ],
        }
        ruta = tmp_path / "multi.xlsx"
        resultado = exportador.exportar_excel_multihoja(datos, ruta)
        assert resultado.exists()

        import openpyxl
        wb = openpyxl.load_workbook(ruta)
        assert "Ventas" in wb.sheetnames
        assert "Compras" in wb.sheetnames
        ws = wb["Ventas"]
        assert ws.cell(1, 1).value == "Numero"
        assert ws.cell(2, 2).value == 605.0
        wb.close()

    def test_libro_diario_excel(self, exportador, asientos_ejemplo, tmp_path):
        ruta = tmp_path / "diario.xlsx"
        resultado = exportador.exportar_libro_diario_excel(asientos_ejemplo, ruta)
        assert resultado.exists()

        import openpyxl
        wb = openpyxl.load_workbook(ruta)
        ws = wb["Libro Diario"]
        assert ws.cell(1, 1).value == "Asiento"
        assert ws.max_row == 7  # 1 cabecera + 6 partidas
        wb.close()


class TestFormatoDecimal:
    def test_float(self, exportador):
        assert exportador._formato_decimal(1234.567) == "1234.57"

    def test_none(self, exportador):
        assert exportador._formato_decimal(None) == "0.00"

    def test_cero(self, exportador):
        assert exportador._formato_decimal(0) == "0.00"
