# -*- coding: utf-8 -*-
"""
Genera gastos_gerardo_2025.xlsx para Quipu a partir de FACTURAS 2025/.
Lee cada PDF con pdfplumber. Para los que no tienen texto (escaneados o
texto espejado) usa Mistral OCR como fallback.
"""

import pdfplumber
import os
import re
import base64
import unicodedata
import openpyxl
import json
import hashlib
from pathlib import Path
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Cache OCR en disco ────────────────────────────────────────────────────────
# Evita llamar a Mistral/GPT en ejecuciones repetidas del mismo PDF.
# Almacena: {sha256_del_pdf: {"texto_ocr": "...", "gpt": {...}}}
CACHE_FILE = Path(__file__).parent / "ocr_cache_gerardo.json"

def _cache_clave(path: str) -> str:
    """SHA256 de los primeros 64 KB del PDF como clave de cache."""
    with open(path, "rb") as f:
        return hashlib.sha256(f.read(65536)).hexdigest()

def _cache_cargar() -> dict:
    if CACHE_FILE.exists():
        try:
            return json.loads(CACHE_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}

def _cache_guardar(cache: dict):
    CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")

# Cache en memoria para la sesión actual
_CACHE = _cache_cargar()

CARPETA  = "c:/Users/carli/PROYECTOS/CONTABILIDAD/clientes/gerardo-gonzalez-callejon/FACTURAS 2025"
TEMPLATE = "c:/Users/carli/Downloads/gastos.xlsx"
SALIDA   = "c:/Users/carli/Downloads/gastos_gerardo_2025.xlsx"

# ── Tabla de proveedores conocidos ────────────────────────────────────────────
# clave norm → (NIF, nombre_oficial, actividad)
PROVEEDORES = {
    'alarma':                       ('A26106013', 'SECURITAS DIRECT ESPANA SAU',   'compartido'),
    'internet':                     ('A82528548', 'XFERA MOVILES SAU',             'compartido'),
    'local':                        ('A08663619', 'CAIXABANK SA',                  'compartido'),
    'letra local':                  ('A08663619', 'CAIXABANK SA',                  'compartido'),
    'autonomos':                    ('Q2827003A', 'TGSS REGIMEN AUTONOMOS',        'compartido'),
    'seguros rg':                   ('A58333261', 'VIDACAIXA SA DE SEGUROS',       'compartido'),
    'seguro':                       ('A58333261', 'VIDACAIXA SA DE SEGUROS',       'compartido'),
    'seguro local mapfre':          ('A28007033', 'MAPFRE ESPANA SA',              'compartido'),
    'rg':                           ('A58333261', 'VIDACAIXA SA DE SEGUROS',       'compartido'),
    'rg eva':                       ('A58333261', 'VIDACAIXA SA DE SEGUROS',       'estetica'),
    'comunidad local':              ('A08663619', 'CAIXABANK SA',                  'compartido'),
    'comunidad':                    ('A08663619', 'CAIXABANK SA',                  'compartido'),
    'comunidad derrama':            ('A08663619', 'CAIXABANK SA',                  'compartido'),
    'datafono':                     ('A08663619', 'CAIXABANK SA',                  'compartido'),
    'ibi local':                    ('A08663619', 'CAIXABANK SA',                  'compartido'),
    'colegio':                      ('Q4100650C', 'COLEGIO PODOLOGOS ANDALUCIA',   'podologia'),
    'web':                          ('25686266G', 'FERNANDO SUSIN MALDONADO',      'compartido'),
    'google':                       ('IE6388047V','GOOGLE IRELAND LTD',            'compartido'),
    'google estetica':              ('IE6388047V','GOOGLE IRELAND LTD',            'estetica'),
    'facebook':                     ('B85834729', 'META PLATFORMS IRELAND LTD',    'estetica'),
    'meta estetica':                ('B85834729', 'META PLATFORMS IRELAND LTD',    'estetica'),
    'clinni':                       ('B91622753', 'VIAFISIO SL',                   'podologia'),
    'clinni estetica':              ('B91622753', 'VIAFISIO SL',                   'estetica'),
    'agenda clinni':                ('B91622753', 'VIAFISIO SL',                   'podologia'),
    'clinni podologia':             ('B91622753', 'VIAFISIO SL',                   'podologia'),
    'electrodo estetica':           ('B97053623', 'CELDUAL SL',                    'estetica'),
    'celdual estetica':             ('B97053623', 'CELDUAL SL',                    'estetica'),
    'skinclinic estetica':          ('B53661690', 'SKINCLINIC SL',                 'estetica'),
    'skinclinic devolucion':        ('B53661690', 'SKINCLINIC SL',                 'estetica'),
    'skeyndor estetica':            ('B63091870', 'SKEYNDOR SL',                   'estetica'),
    'botox estetica':               (None,        'DRA ESTETICA',                  'estetica'),
    'lidocaina estetica':           (None,        'DRA ESTETICA',                  'estetica'),
    'dh material medico estetica':  (None,        'DH MATERIAL MEDICO',            'estetica'),
    'dh material medico':           (None,        'DH MATERIAL MEDICO',            'podologia'),
    'wakeupconsultoria estetica':   ('B87409686', 'WAKEUP CONSULTORIA SL',         'estetica'),
    'wakeup consultoria estetica':  ('B87409686', 'WAKEUP CONSULTORIA SL',         'estetica'),
    'nanosat revision laser estetica': (None,     'NANOSAT REVISION LASER',        'estetica'),
    'quirumed lampara led estetica':('B97267405', 'QUIRUMED SL',                   'estetica'),
    'evocare electrodos estetica':  ('19994601B', 'EVOCARE',                       'estetica'),
    'groupon estetica':             ('B85834729', 'GROUPON SL',                    'estetica'),
    'asesoria laboral':             (None,        'ASESORIA LABORAL',              'compartido'),
    'podologa may':                 (None,        'PODOLOGA MAY',                  'podologia'),
    'podologa marta':               ('44523191K', 'MARTA NUEVALOS ALVAREZ',        'podologia'),
    'psicologia gloria':            ('06285776Z', 'GLORIA SAINERO TIRADO',         'podologia'),
    'psicologa gloria':             ('06285776Z', 'GLORIA SAINERO TIRADO',         'podologia'),
    'coexsur':                      (None,        'COEXSUR',                       'podologia'),
    'podiatech':                    ('B62757695', 'PODIATECH SL',                  'podologia'),
    'podiatech materiales':         ('B62757695', 'PODIATECH SL',                  'podologia'),
    'podiatech materiales taller':  ('B62757695', 'PODIATECH SL',                  'podologia'),
    'podoservice':                  ('B63021943', 'PODOSERVICE SL',                'podologia'),
    'podoservice fieltro':          ('B63021943', 'PODOSERVICE SL',                'podologia'),
    'herbitas':                     ('B63021943', 'HERBITAS SL',                   'podologia'),
    'herbitas forro':               ('B63021943', 'HERBITAS SL',                   'podologia'),
    'fisaude':                      ('B32403495', 'FISAUDE SL',                    'podologia'),
    'fisaude devolucion':           ('B32403495', 'FISAUDE SL',                    'podologia'),
    'fresco cremas':                ('B58618323', 'FRESCO PODOLOGIA SL',           'podologia'),
    'fresco mangos y fresas':       ('B58618323', 'FRESCO PODOLOGIA SL',           'podologia'),
    'proteccion datos':             ('B01664762', 'PROTECCION DATOS SL',           'compartido'),
    'gme apse4 podologia':          (None,        'GME APSE4',                     'podologia'),
    'vectem cremas podologia':      ('A08185431', 'VECTEM SA',                     'podologia'),
    'dra estetica':                 ('03907517R', 'CAROLINA LARA PALMERO',         'estetica'),
    'luz':                          (None,        'ENDESA',                        'compartido'),
    'luz 2t':                       (None,        'ENDESA',                        'compartido'),
    'sumevi':                       (None,        'SUMEVI',                        'podologia'),
    'aquaservice':                  ('A18109200', 'VIVA AQUA SERVICE SPAIN SA',    'compartido'),
    'estores':                      ('B60304391', 'ESTORES SL',                    'compartido'),
    'luces letras fachada':         (None,        'LUCES LETRAS FACHADA',          'compartido'),
    'metacrilato estetica':         (None,        'METACRILATO',                   'estetica'),
    'namrol manguera':              ('B61547881', 'NAMROL SL',                     'podologia'),
    'gestion residuos':             ('76638663',  'GESTION RESIDUOS',              'podologia'),
    'proser clinic alicates':       (None,        'PROSER CLINIC',                 'podologia'),
    'proser clinic hojas bisturi':  (None,        'PROSER CLINIC',                 'podologia'),
    'indemnizacion eva':            (None,        'INDEMNIZACION EVA',             'estetica'),
    'tgss':                         ('Q2827003A', 'TGSS',                          'compartido'),
}

IAE = {
    'podologia':  ('II',  '869'),
    'estetica':   ('I',   '971.2'),
    'compartido': ('',    ''),
}


def norm(s):
    return unicodedata.normalize('NFD', s or '').encode('ascii', 'ignore').decode().lower().strip()


def nombre_proveedor(filename):
    base = os.path.splitext(filename)[0]
    partes = base.split(' ', 2)
    return partes[2].strip() if len(partes) >= 3 else base


def info_proveedor(nombre):
    clave = norm(nombre)
    if clave in PROVEEDORES:
        return PROVEEDORES[clave]
    # Buscar parcial
    for k, v in PROVEEDORES.items():
        if clave.startswith(k) or clave == k:
            return v
    # Buscar por palabras clave del nombre
    for k, v in PROVEEDORES.items():
        if k in clave:
            return v
    if 'estetica' in clave:
        return (None, nombre, 'estetica')
    return (None, nombre, 'podologia')


# ── Parsear número ────────────────────────────────────────────────────────────
def parse_num(s):
    if not s:
        return None
    s = str(s).strip()
    # Quitar símbolo moneda
    s = s.lstrip('€$').rstrip('€$').strip()
    s = s.replace('\xa0', '').replace(' ', '')
    if not s:
        return None
    # Negativo
    negativo = s.startswith('-')
    s = s.lstrip('-')
    # Formato europeo: 1.234,56 o 1.234,6
    if re.match(r'^\d{1,3}(\.\d{3})+,\d{1,2}$', s):
        v = float(s.replace('.', '').replace(',', '.'))
        return -v if negativo else v
    # Simple con coma: 1234,56 o 1234,6
    if re.match(r'^\d+,\d{1,2}$', s):
        v = float(s.replace(',', '.'))
        return -v if negativo else v
    # Simple con punto: 1234.56
    if re.match(r'^\d+\.\d{1,2}$', s):
        v = float(s)
        return -v if negativo else v
    # Entero
    if re.match(r'^\d+$', s):
        v = float(s)
        return -v if negativo else v
    return None


def es_razonable(v):
    return v is not None and abs(v) > 0.01 and abs(v) < 100000


PATRON_IMPORTE = re.compile(
    r'^-?(?:\d{1,3}(?:\.\d{3})*,\d{1,2}|\d+,\d{1,2}|\d+\.\d{1,2})$'
)

def nums_de_words(words):
    """Extrae importes con decimales (evita códigos postales enteros)."""
    resultado = []
    for w in words:
        t = w['text'].strip('€$').strip()
        if PATRON_IMPORTE.match(t):
            v = parse_num(t)
            if v is not None and es_razonable(v):
                resultado.append(v)
    return resultado


# ── Extraer importe del nombre del archivo ────────────────────────────────────
def importe_de_filename(filename):
    """Extrae importe si está codificado en el nombre: 'Groupon Comision 104.53' o 'Devolucion -17,97'."""
    base = os.path.splitext(filename)[0]
    # Devoluciones: buscar '-XX,XX' o '-XX.XX'
    m = re.search(r'[-–]\s*(\d+[.,]\d{2})\b', base)
    if m:
        v = parse_num(m.group(1))
        if v:
            return -v  # es una devolución
    # Groupon: buscar 'Comision XX.XX' o simplemente número al final
    m = re.search(r'[Cc]omisi[oó]n\s+([\d.,]+)', base)
    if m:
        v = parse_num(m.group(1))
        if v and es_razonable(v):
            return v
    return None


# ── Mistral OCR fallback ───────────────────────────────────────────────────────
def _mistral_ocr_texto(path):
    """Llama a Mistral OCR (con cache en disco). $0.002/pág → no repetir llamadas."""
    clave = _cache_clave(path)
    entrada = _CACHE.get(clave, {})
    if "texto_ocr" in entrada:
        print(f"  [OCR-CACHE] {os.path.basename(path)}")
        return entrada["texto_ocr"]

    key = os.environ.get("MISTRAL_API_KEY", "")
    if not key:
        return ""
    try:
        from mistralai import Mistral
        client = Mistral(api_key=key)
        with open(path, "rb") as f:
            pdf_b64 = base64.standard_b64encode(f.read()).decode("utf-8")
        resp = client.ocr.process(
            model="mistral-ocr-latest",
            document={"type": "document_url",
                      "document_url": f"data:application/pdf;base64,{pdf_b64}"},
        )
        texto = ""
        if hasattr(resp, "pages") and resp.pages:
            for p in resp.pages:
                texto += p.markdown + "\n"
        # Guardar en cache
        entrada["texto_ocr"] = texto
        _CACHE[clave] = entrada
        _cache_guardar(_CACHE)
        return texto
    except Exception as e:
        print(f"  [OCR] Error en {os.path.basename(path)}: {e}")
        return ""


def _gpt_parsear_ocr(texto_ocr, path=None):
    """Llama a GPT-4o para extraer campos estructurados (con cache en disco).
    Sólo se usa como segundo paso cuando el regex no encuentra número de factura.
    Devuelve dict con claves: numero_factura, base_imponible, iva_porcentaje,
    irpf_porcentaje, total. Todos pueden ser None.
    """
    # Cache por hash del texto OCR (no del PDF, ya puede estar cacheado)
    clave_gpt = hashlib.sha256(texto_ocr.encode()).hexdigest()[:16]
    if path:
        clave = _cache_clave(path)
        entrada = _CACHE.get(clave, {})
        if "gpt" in entrada:
            print(f"  [GPT-CACHE] {os.path.basename(path)}")
            return entrada["gpt"]

    key = os.environ.get("OPENAI_API_KEY", "")
    if not key:
        return {}
    try:
        from openai import OpenAI
        import json as _json
        client = OpenAI(api_key=key)
        prompt = (
            "Eres un experto en contabilidad española. Analiza el documento y devuelve "
            "SOLO JSON con estos campos (null si no aparecen):\n"
            '{"numero_factura":null,"base_imponible":null,"iva_porcentaje":null,'
            '"irpf_porcentaje":null,"total":null}'
        )
        resp = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": prompt},
                {"role": "user", "content": f"Documento:\n\n{texto_ocr}"},
            ],
            response_format={"type": "json_object"},
            temperature=0.1,
            max_tokens=200,
        )
        datos = _json.loads(resp.choices[0].message.content)
        # Guardar en cache
        if path:
            clave = _cache_clave(path)
            entrada = _CACHE.get(clave, {})
            entrada["gpt"] = datos
            _CACHE[clave] = entrada
            _cache_guardar(_CACHE)
        return datos
    except Exception as e:
        print(f"  [GPT] Error: {e}")
        return {}


def _limpiar_markdown_ocr(texto):
    """Convierte el markdown de tablas OCR en texto plano para facilitar parsing.

    Transforma esto:
      |  TOTAL  |
      | --- |
      |  49,61 Eurs  |
    En: "TOTAL 49,61 EUR"
    """
    # Eliminar filas separadoras de tabla (| --- | | :---: |)
    texto = re.sub(r'\|[\s:|-]+\|\n?', '', texto)
    # Juntar celdas de tabla en una sola línea por fila
    texto = re.sub(r'\|', ' ', texto)
    # Normalizar "Eurs" y "Euros" → "EUR" para que los patrones € funcionen
    texto = re.sub(r'\bEurs?\b', 'EUR', texto, flags=re.IGNORECASE)
    # Colapsar múltiples espacios
    texto = re.sub(r'  +', ' ', texto)
    return texto


# ── Extractores por tipo ──────────────────────────────────────────────────────
def extraer_caixabank_movimiento(texto):
    m = re.search(r'Importe\s+[-\u2013]\s*([\d.,]+)\s*euros', texto, re.IGNORECASE)
    if m:
        v = parse_num(m.group(1))
        if es_razonable(v):
            return v
    return None


def extraer_domiciliacion(words):
    """Domiciliacion SEPA: el importe es el número razonable más grande en words."""
    candidatos = [v for v in nums_de_words(words) if 1 < v < 10000]
    return max(candidatos) if candidatos else None


def extraer_prestamo(texto, words):
    """Recibo de préstamo CaixaBank: busca 'IMPORTE{} ... XXX,XX'."""
    # Texto contiene: "}IMPORTE{} ... FECHA VALOR XXX,XX"
    m = re.search(r'FECHA\s+VALOR[.\s}]*[\d.]+\s+([\d.,]+)', texto, re.IGNORECASE)
    if m:
        v = parse_num(m.group(1))
        if es_razonable(v):
            return v
    # Fallback: max en words
    candidatos = [v for v in nums_de_words(words) if 100 < v < 10000]
    return max(candidatos) if candidatos else None


def extraer_google(texto, words):
    """Google Ads / Google Ireland."""
    # Buscar en texto
    for pat in [
        r'Amount\s+due[:\s]*(?:EUR\s*)?([€]?[\d.,]+)',
        r'Total[:\s]+(?:EUR\s*)?([€]?[\d.,]+)',
        r'Importe\s+total[:\s]+([€]?[\d.,]+)',
    ]:
        m = re.search(pat, texto, re.IGNORECASE)
        if m:
            v = parse_num(m.group(1))
            if es_razonable(v):
                return v
    # Buscar en words: '€XXX.XX' o '€XXX,XX'
    for w in words:
        t = w['text']
        if t.startswith('€') and len(t) > 1:
            v = parse_num(t[1:])
            if es_razonable(v):
                return v
    # Fallback max
    candidatos = [v for v in nums_de_words(words) if es_razonable(v)]
    return max(candidatos) if candidatos else None


def extraer_meta(texto, words):
    """Meta Platforms / Facebook Ads."""
    for pat in [
        r'Importe\s+total\s+facturado\s+([\d.,]+)',
        r'[€\$]\s*([\d.,]+)\s*(?:\(EUR\)|EUR)',
        r'([\d.,]+)\s*[€]\s*(?:\(EUR\))',
    ]:
        m = re.search(pat, texto, re.IGNORECASE)
        if m:
            v = parse_num(m.group(1))
            if es_razonable(v):
                return v
    candidatos = [v for v in nums_de_words(words) if es_razonable(v)]
    return max(candidatos) if candidatos else None


def extraer_clinni(texto):
    base, iva_pct, irpf_pct, total = None, None, None, None
    m = re.search(r'Total:\s*([€$]?[\d.,]+)', texto, re.IGNORECASE)
    if m:
        total = parse_num(m.group(1))
    m = re.search(r'Subtotal:\s*([€$]?[\d.,]+)', texto, re.IGNORECASE)
    if m:
        base = parse_num(m.group(1))
    m = re.search(r'IVA\s*\((\d+)%\)', texto, re.IGNORECASE)
    if m:
        iva_pct = int(m.group(1))
    return base, iva_pct, irpf_pct, total


def extraer_factura_estandar(texto, words):
    """Factura estándar con fallback a max(words)."""
    def buscar_keyword(patrones):
        for p in patrones:
            m = re.search(p, texto, re.IGNORECASE)
            if m:
                v = parse_num(m.group(1))
                if es_razonable(v):
                    return v
        return None

    total = buscar_keyword([
        r'IMPORTE\s+TOTAL\s+FACTURA\s+([€]?[\d.,]+)',
        r'TOTAL\s+FACTURA\s+[^0-9€]*([€]?[\d.,]+)',
        r'Total\s+Debe\s+[^0-9€]*([€]?[\d.,]+)',
        r'Total\s*(?:a\s*pagar)?[:\s]+([€]?[\d.,]+)',
        r'TOTAL\s*[:\s]+([€]?[\d.,]+)',
        r'([€]?[\d.,]+)\s*[€]\s*Total',
        r'([€]?[\d.,]+)\s*[€]\s*$',
    ])

    base = buscar_keyword([
        r'Base\s*[Ii]mponible[:\s]+([€]?[\d.,]+)',
        r'BASE\s+IMPONIBLE[:\s]+([€]?[\d.,]+)',
        r'Subtotal[:\s]+([€]?[\d.,]+)',
        r'Base[:\s]+([€]?[\d.,]+)',
        r'Net[:\s]+([€]?[\d.,]+)',
        r'Base\.[Ii]mp\.[:\s]+([€]?[\d.,]+)',
    ])

    iva_pct = None
    m = re.search(r'IVA[^%0-9]*([0-9]+)\s*%', texto, re.IGNORECASE)
    if m:
        iva_pct = int(m.group(1))

    irpf_pct = None
    m = re.search(r'RETENCI[OÓ]N\s*(\d+)\s*%', texto, re.IGNORECASE)
    if not m:
        m = re.search(r'IRPF[^0-9]*([0-9]+)\s*%', texto, re.IGNORECASE)
    if m:
        irpf_pct = int(m.group(1))

    # Fallback: max de words con decimales (evita coger códigos postales enteros)
    if total is None:
        candidatos = [v for v in nums_de_words(words) if abs(v) >= 1]
        if candidatos:
            total = max(candidatos, key=abs)

    return base, iva_pct, irpf_pct, total


def extraer_nif_texto(texto):
    nifs = re.findall(
        r'\b([A-Z]{1,2}[0-9]{6,7}[A-Z0-9]|[0-9]{8}[A-Z]|IE\s*[0-9A-Z]+)\b',
        texto.upper()
    )
    excluir = {'76638663H', 'ES76638663H', 'A08663619', 'Q2827003A'}
    return [n for n in nifs if n not in excluir]


def extraer_numero_factura(texto):
    def _valido(n):
        """El número de factura debe contener al menos un dígito y tener 2-30 chars."""
        return (2 <= len(n) <= 30
                and not n.isspace()
                and re.search(r'\d', n)
                and not re.match(r'^\d{2}/\d{2}/\d{4}$', n))

    # Excluir NIFs/CIFs (formato fiscal, no número de factura)
    # Los NIFs/CIFs NUNCA contienen '/' — eso ya los distingue de números como I2025/2777
    def _es_nif_cif(n):
        if '/' in n:
            return False
        s = n.replace('-', '')
        return (re.match(r'^[A-Z]\d{7}[A-Z0-9]$', s)   # CIF/NIF empresa
                or re.match(r'^\d{8}[A-Z]$', s))         # NIF persona

    patrones = [
        # Google: Invoice number: 5195726375
        (r'Invoice\s+(?:number|No\.?)\s*[:\s]*([A-Z0-9/\-]{4,30})', 0),
        # Etiquetas explícitas en español (misma línea)
        (r'N[úu·ÚU]m(?:ero)?\s*(?:de\s*)?[Ff]actura\s*[:\s#.]+([A-Z0-9][A-Z0-9/\-]{1,29})', 0),
        (r'[Ff]actura\s+N[úu]mero\s*[:\s]+([A-Z0-9][A-Z0-9/\-]{1,29})', 0),   # Clinni: Factura Número: I2025/xxx
        (r'N[º°]\s*(?:de\s*)?[Ff]actura\s*[:\s]+([A-Z0-9][A-Z0-9/\-]{1,29})', 0),
        (r'[Ff]actura\s+N[º°]\s*[:\s#.]*([A-Z0-9]\d[A-Z0-9/\-]{0,28})', 0),   # Marta: FACTURA Nº: 01
        (r'Fra\.?\s*N[º°]?\s*[:\s]+([A-Z0-9][A-Z0-9/\-]{1,29})', 0),
        # SkinClinic: "NÚM. FACTURA" separado del valor por headers de columna (permite \n)
        (r'N.M\.?\s*FACTURA\b[\s\S]{1,120}?([A-Z]{1,4}-\d{4,9})', re.DOTALL),
        # WakeUp: "Código Factura\nF2025001009"
        (r'C.digo\s+Factura[\s\S]{0,50}?([A-Z]\d{7,15})', re.DOTALL),
        # DH Material: "Número Fecha Cliente A1 2518386 14/05/2025"
        (r'N.mero\s+Fecha\s+Cliente\s+([A-Z]\d+(?:\s+\d+)?)', 0),
        # Skeyndor VN/556213
        (r'\b(V[NA]/\d{5,9})\b', 0),
        # Vectem: "Factura N°\nMALAGA\n23-04-2025 273501" — cruza \n con [\s\S]
        (r'[Ff]actura\s+N.[\s\S]{0,100}?(\d{5,8})\s+\d', re.DOTALL),
    ]
    for pat, flags in patrones:
        m = re.search(pat, texto, re.IGNORECASE | flags)
        if m:
            n = m.group(1).strip().rstrip('.')
            if _valido(n) and not _es_nif_cif(n):
                return n
    return None


# ── Procesador principal ──────────────────────────────────────────────────────
def procesar_pdf(path, filename):
    nombre_prov = nombre_proveedor(filename)
    nif_config, nombre_oficial, actividad = info_proveedor(nombre_prov)
    fecha_str = filename[:8]

    try:
        with pdfplumber.open(path) as pdf:
            texto = ''
            words = []
            for page in pdf.pages:
                texto += page.extract_text() or ''
                words += page.extract_words()
    except Exception as e:
        return _resultado(filename, fecha_str, nombre_prov, nombre_oficial, nif_config, actividad,
                          None, None, None, None, None, True, True, error=str(e))

    sin_texto = not texto.strip() and not words

    # Detectores de tipo
    es_caixabank_mov   = 'Detalle del movimiento' in texto and 'Importe' in texto
    es_domiciliacion   = 'IdentificadorcuentaIBAN' in texto and 'Domiciliaci' in texto
    es_transferencia   = 'Ordendetransferencia' in texto or 'Orden de transferencia' in texto
    es_prestamo        = ('Recibo' in texto and 'OFICINA03889' in texto) or 'Recibo parcial' in texto
    es_mybox           = 'MyBox' in texto or 'Cargo Agrupado' in texto
    es_clinni          = 'Viafisio' in texto or 'Factura de Clinni' in texto
    es_google          = 'Google Ireland' in texto or 'google.com' in texto.lower()
    es_meta_facebook   = ('Meta Platforms' in texto or 'facebook.com' in texto.lower()
                          or 'Clinica G Gonzalez' in texto or 'Merrion Road' in texto)
    es_internet_rot    = (('XFERA' in texto.upper() or 'MASMOVIL' in texto.upper())
                          and 'Periodo facturado' in texto)
    es_aquaservice_rot = 'VIVA AQUA' in texto.upper() or 'AUQA' in texto  # mirrored

    base, iva_pct, irpf_pct, total = None, None, None, None
    num_factura = None
    es_factura = True

    # Intentar extraer importe del nombre del archivo (Groupon, Devoluciones)
    importe_fn = importe_de_filename(filename)

    # OCR fallback: escaneados (sin_texto), texto espejado (es_internet_rot) y
    # PDFs de Internet donde pdfplumber extrae texto ilegible
    es_internet_fn = 'internet' in nombre_prov.lower()
    usar_ocr = (sin_texto or es_internet_rot or es_internet_fn) and importe_fn is None
    if usar_ocr:
        print(f"  [OCR] {filename}")
        texto_ocr = _mistral_ocr_texto(path)
        if texto_ocr.strip():
            texto_limpio = _limpiar_markdown_ocr(texto_ocr)
            base, iva_pct, irpf_pct, total = extraer_factura_estandar(texto_limpio, [])
            num_factura = extraer_numero_factura(texto_limpio)
            # GPT-4o como segundo paso: rellena número y campos que el regex no encontró
            if num_factura is None or total is None:
                print(f"  [GPT] {filename}")
                gpt = _gpt_parsear_ocr(texto_ocr, path=path)
                if num_factura is None:
                    num_factura = gpt.get('numero_factura') or None
                def _to_float(v):
                    if v is None:
                        return None
                    if isinstance(v, (int, float)):
                        return float(v)
                    return parse_num(str(v))

                if total is None and gpt.get('total') is not None:
                    total = _to_float(gpt['total'])
                if base is None and gpt.get('base_imponible') is not None:
                    base = _to_float(gpt['base_imponible'])
                if iva_pct is None and gpt.get('iva_porcentaje') is not None:
                    try:
                        iva_pct = int(float(str(gpt['iva_porcentaje']).replace(',', '.')))
                    except (ValueError, TypeError):
                        pass
                if irpf_pct is None and gpt.get('irpf_porcentaje') is not None:
                    try:
                        irpf_pct = int(float(str(gpt['irpf_porcentaje']).replace(',', '.')))
                    except (ValueError, TypeError):
                        pass

    elif es_caixabank_mov:
        total = extraer_caixabank_movimiento(texto)
        es_factura = False

    elif es_domiciliacion or es_transferencia:
        total = extraer_domiciliacion(words)
        es_factura = False

    elif es_prestamo:
        total = extraer_prestamo(texto, words)
        es_factura = False

    elif es_mybox:
        total = extraer_domiciliacion(words)
        es_factura = False

    elif es_clinni:
        base, iva_pct, irpf_pct, total = extraer_clinni(texto)
        num_factura = extraer_numero_factura(texto)

    elif es_google:
        total = extraer_google(texto, words)
        num_factura = extraer_numero_factura(texto)

    elif es_meta_facebook:
        total = extraer_meta(texto, words)
        num_factura = extraer_numero_factura(texto)

    elif es_aquaservice_rot:
        # La parte legible tiene IMPORTE TOTAL FACTURA
        _, iva_pct, _, total = extraer_factura_estandar(texto, words)
        num_factura = extraer_numero_factura(texto)

    else:
        base, iva_pct, irpf_pct, total = extraer_factura_estandar(texto, words)
        num_factura = extraer_numero_factura(texto)

    # Importe de nombre de archivo como override si no encontramos datos
    if importe_fn is not None and total is None:
        total = importe_fn

    # NIF del proveedor
    nif = nif_config
    if not nif and texto:
        nifs = extraer_nif_texto(texto)
        if nifs:
            nif = nifs[0]

    # Calcular base desde total si no tenemos base
    if base is None and total is not None:
        if iva_pct:
            base = round(total / (1 + iva_pct / 100), 2)
        else:
            base = total

    tiene_datos = total is not None or base is not None

    return _resultado(filename, fecha_str, nombre_prov, nombre_oficial, nif, actividad,
                      base, iva_pct, irpf_pct, total, num_factura, es_factura, sin_texto)


def _resultado(filename, fecha, proveedor, nombre_oficial, nif, actividad,
               base, iva_pct, irpf_pct, total, num_factura, es_factura, sin_texto, error=None):
    tiene_datos = total is not None or base is not None
    return {
        'archivo':       filename,
        'fecha':         fecha,
        'proveedor':     proveedor,
        'nombre_oficial': nombre_oficial,
        'nif':           nif,
        'actividad':     actividad,
        'base':          base,
        'iva_pct':       iva_pct,
        'irpf_pct':      irpf_pct,
        'total':         total,
        'num_factura':   num_factura,
        'es_factura':    es_factura,
        'sin_texto':     sin_texto,
        'tiene_datos':   tiene_datos,
        'error':         error,
    }


# ── Generar Excel ─────────────────────────────────────────────────────────────
def generar_excel(registros):
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb['Sheet 1']
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    fill_pod  = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    fill_est  = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    fill_comp = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    fill_rojo = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')

    fila = 3
    for r in registros:
        actividad   = r['actividad']
        tiene_datos = r['tiene_datos']
        base        = r['base']
        total       = r['total']
        iva_pct     = r['iva_pct']
        irpf_pct    = r['irpf_pct']

        if not tiene_datos:
            fill     = fill_rojo
            base_val = 0.0
        else:
            base_val = base if base is not None else (total or 0)
            fill     = fill_pod if actividad == 'podologia' else (
                       fill_est if actividad == 'estetica' else fill_comp)

        fecha_dt = None
        if r['fecha'] and len(r['fecha']) == 8:
            try:
                fecha_dt = datetime(int(r['fecha'][:4]), int(r['fecha'][4:6]), int(r['fecha'][6:8]))
            except Exception:
                pass

        tipo_doc    = 'Factura' if r['es_factura'] else 'Factura Simplificada'
        nif         = r['nif'] or None
        tipo_fiscal = 'NIF' if nif else 'none'
        iae         = IAE[actividad]

        valores = [
            tipo_doc,
            fecha_dt,
            None,
            fecha_dt,
            'Pagado',
            r['num_factura'],
            nif,
            r['nombre_oficial'],
            None, None, None, None, 'ES', None, None, None, None,
            r['proveedor'],
            round(float(base_val), 2),
            1,
            iva_pct  if iva_pct  is not None else 0,
            0,
            irpf_pct if irpf_pct is not None else 0,
            None,
            iae[0],
            iae[1],
            tipo_fiscal,
        ]

        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=fila, column=col)
            cell.value = val
            cell.fill = fill
            if col in (2, 3, 4) and val is not None:
                cell.number_format = 'DD/MM/YYYY'

        fila += 1

    anchos = [22, 12, 12, 12, 10, 18, 15, 30, 20, 25, 15, 8, 5, 12, 24, 12, 15, 40, 12, 8, 8, 8, 8, 15, 8, 10, 12]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.freeze_panes = 'A3'
    wb.save(SALIDA)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    pdfs = sorted(f for f in os.listdir(CARPETA) if f.lower().endswith('.pdf'))
    print(f"PDFs: {len(pdfs)}")

    registros   = []
    sin_datos   = []

    for f in pdfs:
        r = procesar_pdf(os.path.join(CARPETA, f), f)
        registros.append(r)
        if not r['tiene_datos']:
            sin_datos.append(f)

    generar_excel(registros)

    suma = sum(abs(r['total'] or r['base'] or 0) for r in registros if r['tiene_datos'])
    print(f"Total filas:         {len(registros)}")
    print(f"Con datos:           {len(registros) - len(sin_datos)}")
    print(f"Sin importe (rojo):  {len(sin_datos)}")
    print(f"Suma importes:       {suma:,.2f} EUR")
    print(f"\nSin importe:")
    for f in sin_datos:
        print(f"  {f}")
    print(f"\nGuardado en: {SALIDA}")


if __name__ == '__main__':
    main()
