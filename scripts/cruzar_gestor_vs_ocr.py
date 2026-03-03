"""
Cruce: Libros Gestor (gestoría) vs Facturas OCR (Quipu)
Gerardo Gonzalez Callejon 2025

Detecta:
  A) En gestor Y en OCR         -> todo OK
  B) Solo en gestor (sin OCR)   -> contabilizado pero sin documento digitalizado
  C) Solo en OCR (sin gestor)   -> digitalizado pero aun no contabilizado

Salida: Downloads/cruce_gestor_ocr_gerardo_2025.xlsx
"""
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
import pandas as pd
from difflib import SequenceMatcher

DIR_LIBROS = Path("C:/Users/carli/Downloads/libros_gerardo")
EXCEL_OCR  = Path("C:/Users/carli/Downloads/gastos_gerardo_2025.xlsx")
OUTPUT     = Path("C:/Users/carli/Downloads/cruce_gestor_ocr_gerardo_2025.xlsx")


# ─────────────────────────── helpers ────────────────────────────── #

def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s).upper())
    s = s.encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _empareja(a: str, b: str) -> bool:
    if SequenceMatcher(None, _norm(a), _norm(b)).ratio() >= 0.60:
        return True
    pals_a = {p for p in _norm(a).split() if len(p) > 4}
    pals_b = {p for p in _norm(b).split() if len(p) > 4}
    return bool(pals_a & pals_b)


def _fecha(v) -> date | None:
    if isinstance(v, date):
        return v
    try:
        return pd.to_datetime(v, errors="coerce").date()
    except Exception:
        return None


# ──────────────────── 1. Leer libro principal gestor ────────────── #
# Columnas relevantes (0-indexed tras skiprows=2):
#   0=Ejercicio  1=Periodo  3=Tipo  4=EpigIAE
#   8=FechaExp  10=NumFactura  15=NIF  18=Proveedor
#   25=Total  26=Base  27=TipoIVA  28=CuotaIVA
#   29=CuotaDeducible  32=FechaPago  33=ImportePago  37=RetIRPF

def leer_libro_gestor() -> pd.DataFrame:
    df_raw = pd.read_excel(
        DIR_LIBROS / "Libro de Compras y Gastos 2025.xlsx",
        header=None, skiprows=2
    )

    COLS = {
        1:  "periodo",
        4:  "epigrafe_iae",
        8:  "fecha_exp",
        10: "num_factura",
        15: "nif",
        18: "proveedor",
        25: "total",
        26: "base",
        27: "tipo_iva",
        28: "cuota_iva",
        29: "cuota_deducible",
        37: "retencion_irpf",
    }

    df = df_raw[[c for c in COLS if c < df_raw.shape[1]]].copy()
    df.columns = [COLS[c] for c in COLS if c < df_raw.shape[1]]

    df["fecha_exp"] = pd.to_datetime(df["fecha_exp"], errors="coerce").dt.date
    for col in ["total", "base", "cuota_iva", "cuota_deducible", "retencion_irpf"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Filtrar solo filas con proveedor valido y total != 0
    df = df[df["proveedor"].notna() & (df["total"] != 0)].reset_index(drop=True)

    # Añadir actividad desde libros de actividad (estetica / podologia)
    nums_e  = _nums_actividad("Libro de gastos Estetica 2025.xlsx")
    nums_po = _nums_actividad("Libro de gastos Podologia 2025.xlsx")

    def _actividad(num_fac):
        n = str(num_fac).strip()
        if n in nums_e and n in nums_po:
            return "Ambas"
        if n in nums_e:
            return "Estetica"
        if n in nums_po:
            return "Podologia"
        return "Comun/No asignada"

    df["actividad"] = df["num_factura"].apply(
        lambda v: _actividad(v) if pd.notna(v) else "Comun/No asignada"
    )
    return df


def _nums_actividad(nombre_archivo: str) -> set:
    """Devuelve set de num_factura que aparecen en un libro de actividad."""
    df = pd.read_excel(DIR_LIBROS / nombre_archivo, header=6)
    col_num = df.columns[4]  # posicion 4 = N Factura
    return {str(v).strip() for v in df[col_num].dropna()
            if str(v) not in ("nan", "", "None")}


# ──────────────────── 2. Leer facturas OCR ──────────────────────── #

def leer_ocr() -> pd.DataFrame:
    df = pd.read_excel(EXCEL_OCR, header=1)

    nuevos = []
    seen = {}
    for c in df.columns:
        n = unicodedata.normalize("NFKD", str(c))
        n = n.encode("ascii", "ignore").decode("ascii")
        n = re.sub(r"\s+", "_", n.strip().lower())
        if n in seen:
            seen[n] += 1
            n = f"{n}_{seen[n]}"
        else:
            seen[n] = 0
        nuevos.append(n)
    df.columns = nuevos

    for col in ["fecha_de_emision", "fecha_de_pago"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.date

    for col in ["base_unitaria", "cantidad", "iva_(%)", "retencion_(%)"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    qty  = df.get("cantidad",    pd.Series([1]*len(df))).fillna(1).replace(0, 1)
    base = df.get("base_unitaria", pd.Series([0]*len(df))).fillna(0)
    iva  = df.get("iva_(%)",       pd.Series([0]*len(df))).fillna(0)
    ret  = df.get("retencion_(%)", pd.Series([0]*len(df))).fillna(0)

    df["total_iva"] = base * qty * (1 + iva / 100)
    df["pago_neto"] = df["total_iva"] - base * qty * ret / 100
    return df


# ──────────────────── 3. Cruce ──────────────────────────────────── #

def cruzar(gestor: pd.DataFrame, ocr: pd.DataFrame) -> pd.DataFrame:
    DIAS = 10
    PCT  = 0.08

    col_fecha = "fecha_de_pago" if "fecha_de_pago" in ocr.columns else "fecha_de_emision"
    col_nombre = "nombre" if "nombre" in ocr.columns else None
    col_num    = "numeracion" if "numeracion" in ocr.columns else None

    ocr_filas  = ocr[ocr[col_fecha].notna()].to_dict("records")
    usadas_ocr = set()
    resultados = []

    for _, g in gestor.iterrows():
        f_g   = _fecha(g.get("fecha_exp"))
        total = float(g.get("total") or 0)
        prov  = str(g.get("proveedor") or "")
        num_g = str(g.get("num_factura") or "").strip()

        mejor_idx   = None
        mejor_score = 0
        mejor_fac   = None

        for idx, o in enumerate(ocr_filas):
            if idx in usadas_ocr:
                continue

            f_o    = _fecha(o.get(col_fecha))
            p_neto = float(o.get("pago_neto") or 0)
            nombre = str(o.get(col_nombre, "") or "") if col_nombre else ""
            num_o  = str(o.get(col_num, "") or "").strip()

            if not f_o or not f_g or p_neto <= 0 or total <= 0:
                continue

            diff_d = abs((f_g - f_o).days)
            if diff_d > DIAS:
                continue

            pct = abs(total - p_neto) / max(p_neto, 0.01)
            if pct > PCT:
                continue

            score = (
                max(0, (DIAS - diff_d) / DIAS) * 40
                + max(0, (PCT - pct) / PCT) * 40
            )
            if _empareja(prov, nombre):
                score += 25
            # Num factura exacto
            if num_g and num_o and (
                _norm(num_g) == _norm(num_o)
                or num_g in num_o or num_o in num_g
            ):
                score += 20

            if score > mejor_score:
                mejor_score = score
                mejor_fac   = o
                mejor_idx   = idx

        if mejor_fac:
            usadas_ocr.add(mejor_idx)
            estado = "OK - En gestor y OCR"
        else:
            estado = "Solo en gestor (sin OCR)"

        resultados.append({
            "Estado":            estado,
            "Actividad":         g.get("actividad", ""),
            "Periodo":           g.get("periodo", ""),
            "Fecha gestor":      f_g,
            "Num factura":       num_g,
            "Proveedor gestor":  prov,
            "NIF":               str(g.get("nif", "") or ""),
            "Total gestor":      round(total, 2),
            "Base gestor":       round(float(g.get("base") or 0), 2),
            "Cuota IVA gestor":  round(float(g.get("cuota_iva") or 0), 2),
            "IVA Deducible":     round(float(g.get("cuota_deducible") or 0), 2),
            "Retencion IRPF":    round(float(g.get("retencion_irpf") or 0), 2),
            # OCR
            "Fecha OCR":         _fecha(mejor_fac.get(col_fecha)) if mejor_fac else None,
            "Num factura OCR":   mejor_fac.get(col_num, "") if mejor_fac and col_num else "",
            "Proveedor OCR":     mejor_fac.get(col_nombre, "") if mejor_fac and col_nombre else "",
            "Total OCR":         round(float(mejor_fac.get("pago_neto") or 0), 2) if mejor_fac else None,
            "Score match":       round(mejor_score, 1),
        })

    # Facturas OCR sin registro en gestor
    for idx, o in enumerate(ocr_filas):
        if idx in usadas_ocr:
            continue
        nombre = str(o.get(col_nombre, "") or "") if col_nombre else ""
        num_o  = str(o.get(col_num, "") or "").strip()
        resultados.append({
            "Estado":            "Solo en OCR (sin gestor)",
            "Actividad":         "",
            "Periodo":           "",
            "Fecha gestor":      None,
            "Num factura":       "",
            "Proveedor gestor":  "",
            "NIF":               "",
            "Total gestor":      None,
            "Base gestor":       None,
            "Cuota IVA gestor":  None,
            "IVA Deducible":     None,
            "Retencion IRPF":    None,
            "Fecha OCR":         _fecha(o.get(col_fecha)),
            "Num factura OCR":   num_o,
            "Proveedor OCR":     nombre,
            "Total OCR":         round(float(o.get("pago_neto") or 0), 2),
            "Score match":       None,
        })

    return pd.DataFrame(resultados)


# ──────────────────── 4. Exportar ───────────────────────────────── #

def exportar(df: pd.DataFrame, gestor: pd.DataFrame, filepath: Path):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    VERDE   = PatternFill("solid", fgColor="C6EFCE")   # OK
    AZUL    = PatternFill("solid", fgColor="BDD7EE")   # solo gestor
    NARANJA = PatternFill("solid", fgColor="FFEB9C")   # solo OCR
    CAB_BG  = PatternFill("solid", fgColor="17375E")
    CAB_FT  = Font(bold=True, color="FFFFFF")
    BD      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _cab(ws, cols):
        for ci, c in enumerate(cols, 1):
            cell = ws.cell(1, ci, c)
            cell.fill = CAB_BG; cell.font = CAB_FT
            cell.alignment = Alignment(horizontal="center"); cell.border = BD

    def _color(estado):
        if "OK" in estado:   return VERDE
        if "OCR" in estado:  return NARANJA
        return AZUL

    # ---- Cruce completo ----
    ws = wb.active
    ws.title = "Cruce Completo"
    cols = list(df.columns)
    _cab(ws, cols)
    for ri, row in df.iterrows():
        bg = _color(row["Estado"])
        for ci, col in enumerate(cols, 1):
            val = row[col]
            if hasattr(val, "date"):
                val = val.date()
            elif val != val:
                val = None
            c = ws.cell(ri + 2, ci, val)
            c.border = BD; c.fill = bg
    anchos = {"Estado": 28, "Proveedor gestor": 35, "Proveedor OCR": 35,
              "Actividad": 12, "Periodo": 8}
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = anchos.get(col, 15)
    ws.freeze_panes = "A2"

    # ---- Solo en gestor (sin OCR) ----
    ws2 = wb.create_sheet("Sin documento OCR")
    df2 = df[df["Estado"] == "Solo en gestor (sin OCR)"].sort_values("Fecha gestor")
    c2  = ["Actividad", "Periodo", "Fecha gestor", "Num factura",
           "Proveedor gestor", "NIF", "Total gestor", "Base gestor",
           "Cuota IVA gestor", "Retencion IRPF"]
    _cab(ws2, c2)
    for ri2, (_, row) in enumerate(df2.iterrows(), 2):
        for ci, col in enumerate(c2, 1):
            val = row[col]
            if hasattr(val, "date"): val = val.date()
            elif val != val: val = None
            c = ws2.cell(ri2, ci, val)
            c.border = BD; c.fill = AZUL
    ws2.column_dimensions["E"].width = 35; ws2.freeze_panes = "A2"

    # ---- Solo en OCR (sin gestor) ----
    ws3 = wb.create_sheet("Sin contabilizar (OCR)")
    df3 = df[df["Estado"] == "Solo en OCR (sin gestor)"].sort_values("Fecha OCR")
    c3  = ["Fecha OCR", "Num factura OCR", "Proveedor OCR", "Total OCR"]
    _cab(ws3, c3)
    for ri3, (_, row) in enumerate(df3.iterrows(), 2):
        for ci, col in enumerate(c3, 1):
            val = row[col]
            if hasattr(val, "date"): val = val.date()
            elif val != val: val = None
            c = ws3.cell(ri3, ci, val)
            c.border = BD; c.fill = NARANJA
    ws3.column_dimensions["C"].width = 35; ws3.freeze_panes = "A2"

    # ---- Resumen por proveedor ----
    ws4 = wb.create_sheet("Resumen Proveedores")
    r4 = (
        df[df["Total gestor"].notna()]
        .groupby(["Proveedor gestor", "Actividad", "Estado"])
        .agg(N=("Total gestor", "count"), Total=("Total gestor", "sum"))
        .reset_index()
        .sort_values("Total", ascending=False)
    )
    r4["Total"] = r4["Total"].round(2)
    _cab(ws4, list(r4.columns))
    for ri4, row in enumerate(r4.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws4.cell(ri4, ci, val).border = BD
    ws4.column_dimensions["A"].width = 35; ws4.freeze_panes = "A2"

    # ---- Estadisticas ----
    ws5 = wb.create_sheet("Estadisticas")
    ok  = df[df["Estado"].str.startswith("OK")]
    sg  = df[df["Estado"] == "Solo en gestor (sin OCR)"]
    so  = df[df["Estado"] == "Solo en OCR (sin gestor)"]

    stats = [
        ["CRUCE GESTOR vs OCR -- GERARDO GONZALEZ 2025", ""],
        ["", ""],
        ["LIBROS GESTOR (Libro de Compras y Gastos 2025)", ""],
        ["  Total gastos registrados", len(gestor)],
        ["  Total EUR registrado", round(float(gestor["total"].sum()), 2)],
        ["  De actividad Estetica",
         len(gestor[gestor["actividad"] == "Estetica"])],
        ["  De actividad Podologia",
         len(gestor[gestor["actividad"] == "Podologia"])],
        ["  Comunes o no asignadas",
         len(gestor[~gestor["actividad"].isin(["Estetica", "Podologia"])])],
        ["", ""],
        ["FACTURAS OCR (Quipu)", ""],
        ["  Total facturas", len(ocr if False else df[df["Proveedor OCR"] != ""])],
        ["", ""],
        ["RESULTADO CRUCE", ""],
        ["  OK - En gestor Y en OCR",           len(ok)],
        ["    Total EUR conciliado",             round(float(ok["Total gestor"].sum()), 2)],
        ["", ""],
        ["  Solo en gestor (SIN documento OCR)", len(sg)],
        ["    Total EUR sin documento",          round(float(sg["Total gestor"].sum()), 2)],
        ["", ""],
        ["  Solo en OCR (NO contabilizado aun)", len(so)],
        ["    Total EUR sin contabilizar",       round(float(so["Total OCR"].sum()), 2)],
        ["", ""],
        ["% Gastos gestor con documento OCR",
         f"{100*len(ok)/max(len(gestor),1):.1f}%"],
        ["% Facturas OCR contabilizadas",
         f"{100*len(ok)/max(len(ok)+len(so),1):.1f}%"],
    ]
    for row in stats:
        ws5.append(row)
    ws5["A1"].font = Font(bold=True, size=13)
    ws5.column_dimensions["A"].width = 48
    ws5.column_dimensions["B"].width = 22

    wb.save(filepath)


# ──────────────────── MAIN ──────────────────────────────────────── #

def main():
    print("=== Cruce Gestor vs OCR -- Gerardo Gonzalez ===")
    print()

    print("Cargando libro del gestor...")
    gestor = leer_libro_gestor()
    print(f"  {len(gestor)} gastos en libro gestor | {gestor['total'].sum():.2f} EUR")
    for act in sorted(gestor["actividad"].unique()):
        sub = gestor[gestor["actividad"] == act]
        print(f"    {act}: {len(sub)} facturas | {sub['total'].sum():.2f} EUR")
    print()

    print("Cargando facturas OCR...")
    global ocr
    ocr = leer_ocr()
    print(f"  {len(ocr)} facturas OCR | {ocr['pago_neto'].sum():.2f} EUR")
    print()

    print("Cruzando...")
    resultado = cruzar(gestor, ocr)

    ok = resultado[resultado["Estado"].str.startswith("OK")]
    sg = resultado[resultado["Estado"] == "Solo en gestor (sin OCR)"]
    so = resultado[resultado["Estado"] == "Solo en OCR (sin gestor)"]

    print(f"  OK (en gestor y OCR):          {len(ok):3d}  |  {ok['Total gestor'].sum():.2f} EUR")
    print(f"  Solo en gestor (sin doc OCR):  {len(sg):3d}  |  {sg['Total gestor'].sum():.2f} EUR")
    print(f"  Solo en OCR (sin contabilizar):{len(so):3d}  |  {so['Total OCR'].sum():.2f} EUR")

    if not sg.empty:
        print()
        print("  -- Gastos contabilizados SIN factura OCR --")
        for _, r in sg.sort_values("Fecha gestor").iterrows():
            print(f"    {str(r['Fecha gestor']):<12}  {r['Total gestor']:>9.2f} EUR"
                  f"  {str(r['Actividad']):<12}  {str(r['Proveedor gestor'])[:40]}"
                  f"  ({r['Num factura']})")

    if not so.empty:
        print()
        print("  -- Facturas OCR NO contabilizadas en gestor --")
        for _, r in so.sort_values("Fecha OCR").iterrows():
            imp = r["Total OCR"] or 0
            print(f"    {str(r['Fecha OCR']):<12}  {imp:>9.2f} EUR"
                  f"  {str(r['Proveedor OCR'])[:40]}"
                  f"  ({r['Num factura OCR']})")

    print()
    exportar(resultado, gestor, OUTPUT)
    print(f"Excel guardado: {OUTPUT}")


if __name__ == "__main__":
    main()
