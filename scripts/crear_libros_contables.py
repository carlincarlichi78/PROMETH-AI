"""
Script para generar el Excel de libros contables consultando la API de FacturaScripts.
Genera un archivo .xlsx con 9 pestanas: Ingresos, Gastos, Bienes Inversion,
Registro Fact. Emitidas, Registro Fact. Recibidas, Resumen Trimestral,
Conciliacion Bancaria, Diario Contable, Balance Sumas y Saldos.

Uso: python crear_libros_contables.py <ruta_salida> [--empresa N]
"""

import argparse
import json
import os
import sys
from collections import defaultdict

import openpyxl
import requests
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- Configuracion API ---

BASE_URL = "https://contabilidad.lemonfresh-tuc.com/api/3/"
TOKEN_FALLBACK = "iOXmrA1Bbn8RDWXLv91L"


def obtener_token():
    """Obtiene el token de la variable de entorno o usa el fallback."""
    return os.environ.get("FS_API_TOKEN", TOKEN_FALLBACK)


def api_get(endpoint, params=None):
    """Consulta GET a la API de FacturaScripts con paginacion automatica."""
    token = obtener_token()
    headers = {"Token": token}
    todos = []
    offset = 0
    limite = 50

    while True:
        parametros = {"limit": limite, "offset": offset}
        if params:
            parametros.update(params)
        try:
            resp = requests.get(
                f"{BASE_URL}{endpoint}",
                headers=headers,
                params=parametros,
                timeout=30,
            )
            resp.raise_for_status()
        except requests.RequestException as e:
            print(f"Error consultando {endpoint}: {e}")
            sys.exit(1)

        datos = resp.json()
        if not datos:
            break
        todos.extend(datos)
        if len(datos) < limite:
            break
        offset += limite

    return todos


# --- Clasificacion y utilidades ---


def calcular_trimestre(fecha_str):
    """Calcula el trimestre a partir de una fecha dd-mm-yyyy."""
    partes = fecha_str.split("-")
    mes = int(partes[1])
    if mes <= 3:
        return "T1"
    elif mes <= 6:
        return "T2"
    elif mes <= 9:
        return "T3"
    else:
        return "T4"


def clasificar_gasto(nombre_proveedor):
    """Clasifica el tipo de gasto segun el nombre del proveedor."""
    nombre = (nombre_proveedor or "").upper()
    clasificaciones = {
        "PRIMAFRIO": "Transporte nacional",
        "LOGINET": "Flete maritimo",
        "LNET": "Flete maritimo",
        "PRIMATRANSIT": "Despacho aduanero",
        "TRANSITAINER": "Despacho aduanero",
        "MAERSK": "Naviera",
        "CAUQUEN": "Compra mercaderia",
        "ODOO": "Software",
        "COPYRAP": "Publicidad",
        "PUBLICIDAD": "Publicidad",
        "EL CORTE INGLES": "Material oficina",
        "CORTE INGLES": "Material oficina",
    }
    for clave, tipo in clasificaciones.items():
        if clave in nombre:
            return tipo
    return "Otros"


def es_nota_credito(factura):
    """Determina si una factura es nota de credito (serie R)."""
    return factura.get("codserie", "") == "R"


def signo_factura(factura):
    """Devuelve -1 para notas de credito, 1 para facturas normales."""
    return -1 if es_nota_credito(factura) else 1


def parsear_fecha(fecha_str):
    """Convierte dd-mm-yyyy en fecha legible."""
    return fecha_str


def concepto_factura(factura, lineas):
    """Extrae el concepto de observaciones o de la primera linea."""
    obs = factura.get("observaciones", "")
    if obs and obs.strip():
        return obs.strip()
    lineas_factura = [l for l in lineas if l["idfactura"] == factura["idfactura"]]
    if lineas_factura:
        return lineas_factura[0].get("descripcion", "")
    return ""


# --- Estilos Excel ---


def estilo_cabecera():
    """Retorna diccionario con estilos para cabeceras."""
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill(
            start_color="2F5496", end_color="2F5496", fill_type="solid"
        ),
        "alignment": Alignment(
            horizontal="center", vertical="center", wrap_text=True
        ),
        "border": Border(
            bottom=Side(style="thin"),
            top=Side(style="thin"),
            left=Side(style="thin"),
            right=Side(style="thin"),
        ),
    }


def estilo_moneda():
    """Retorna formato numerico para moneda."""
    return '#,##0.00'


def aplicar_cabecera(ws, cabeceras, anchos):
    """Aplica estilo de cabecera a la primera fila."""
    estilo = estilo_cabecera()
    for col, (titulo, ancho) in enumerate(zip(cabeceras, anchos), 1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.font = estilo["font"]
        celda.fill = estilo["fill"]
        celda.alignment = estilo["alignment"]
        celda.border = estilo["border"]
        ws.column_dimensions[get_column_letter(col)].width = ancho
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cabeceras))}1"
    ws.freeze_panes = "A2"


def formatear_moneda(ws, fila, columnas):
    """Aplica formato moneda a columnas especificas de una fila."""
    fmt = estilo_moneda()
    for col in columnas:
        celda = ws.cell(row=fila, column=col)
        if celda.value is not None:
            celda.number_format = fmt


# --- Carga de datos desde API ---


def convertir_facturas_a_eur(facturas):
    """Convierte campos monetarios de facturas en divisa extranjera a EUR.

    Guarda los valores originales en campos _original para referencia.
    """
    campos_monetarios = ["neto", "totaliva", "totalirpf", "total"]
    n_convertidas = 0
    for f in facturas:
        divisa = f.get("coddivisa", "EUR")
        tc = float(f.get("tasaconv", 1) or 1)
        f["_divisa_original"] = divisa
        f["_tc"] = tc
        if divisa != "EUR" and tc > 0 and tc != 1:
            for campo in campos_monetarios:
                valor_original = float(f.get(campo, 0))
                f[f"_{campo}_original"] = valor_original
                f[campo] = round(valor_original / tc, 2)
            n_convertidas += 1
        else:
            for campo in campos_monetarios:
                f[f"_{campo}_original"] = float(f.get(campo, 0))
    return facturas, n_convertidas


def cargar_datos(idempresa):
    """Carga todos los datos necesarios de la API."""
    print(f"Consultando API para empresa {idempresa}...")

    # Filtro por empresa
    filtro = {"idempresa": idempresa}

    facturas_cliente = api_get("facturaclientes", filtro)
    facturas_proveedor = api_get("facturaproveedores", filtro)
    lineas_cliente = api_get("lineafacturaclientes")
    lineas_proveedor = api_get("lineafacturaproveedores")
    asientos = api_get("asientos", filtro)
    partidas = api_get("partidas")
    subcuentas = api_get("subcuentas")

    # Post-filtro por idempresa (la API FS ignora el filtro en algunos endpoints)
    facturas_cliente = [f for f in facturas_cliente if str(f.get("idempresa")) == str(idempresa)]
    facturas_proveedor = [f for f in facturas_proveedor if str(f.get("idempresa")) == str(idempresa)]
    asientos = [a for a in asientos if str(a.get("idempresa")) == str(idempresa)]

    # Filtrar lineas por facturas de esta empresa
    ids_fc = {f["idfactura"] for f in facturas_cliente}
    ids_fp = {f["idfactura"] for f in facturas_proveedor}
    lineas_cliente = [l for l in lineas_cliente if l["idfactura"] in ids_fc]
    lineas_proveedor = [l for l in lineas_proveedor if l["idfactura"] in ids_fp]

    # Filtrar partidas por asientos de esta empresa
    ids_asientos = {a["idasiento"] for a in asientos}
    partidas = [p for p in partidas if p["idasiento"] in ids_asientos]

    # Mapa de subcuentas para descripciones
    mapa_subcuentas = {s["codsubcuenta"]: s["descripcion"] for s in subcuentas}

    # Convertir facturas en divisa extranjera a EUR
    facturas_cliente, n_cli = convertir_facturas_a_eur(facturas_cliente)
    facturas_proveedor, n_prov = convertir_facturas_a_eur(facturas_proveedor)

    # Convertir lineas de facturas en divisa extranjera a EUR
    tc_por_factura = {}
    for f in facturas_cliente + facturas_proveedor:
        if f.get("_divisa_original", "EUR") != "EUR":
            tc_por_factura[f["idfactura"]] = f["_tc"]

    for linea in lineas_cliente + lineas_proveedor:
        tc = tc_por_factura.get(linea["idfactura"])
        if tc and tc > 0 and tc != 1:
            for campo in ["pvptotal", "pvpunitario"]:
                if campo in linea:
                    linea[campo] = round(float(linea[campo]) / tc, 2)

    print(
        f"  Facturas cliente: {len(facturas_cliente)}, "
        f"Facturas proveedor: {len(facturas_proveedor)}"
    )
    if n_cli + n_prov > 0:
        print(f"  Convertidas a EUR: {n_cli} cliente + {n_prov} proveedor")
    print(f"  Asientos: {len(asientos)}, Partidas: {len(partidas)}")

    return {
        "facturas_cliente": sorted(facturas_cliente, key=lambda f: f["fecha"]),
        "facturas_proveedor": sorted(facturas_proveedor, key=lambda f: f["fecha"]),
        "lineas_cliente": lineas_cliente,
        "lineas_proveedor": lineas_proveedor,
        "asientos": sorted(asientos, key=lambda a: (a["fecha"], a["idasiento"])),
        "partidas": partidas,
        "mapa_subcuentas": mapa_subcuentas,
    }


# --- Generacion de pestanas ---


def crear_libro_ingresos(wb, datos):
    """Pestana 1: Ingresos (facturas de cliente)."""
    ws = wb.create_sheet("Ingresos")
    cabeceras = [
        "Fecha", "Trimestre", "Actividad", "N Factura", "Cliente",
        "CIF Cliente", "Concepto", "Base Imponible", "% IVA",
        "Cuota IVA", "% IRPF", "Retencion IRPF", "Total Factura",
        "Divisa", "TC", "Total Original",
        "Cobrada", "Fecha Cobro", "Forma Pago",
        "Observaciones",
    ]
    anchos = [
        12, 10, 20, 16, 28, 15, 30, 16, 8, 14, 8, 16, 14,
        8, 8, 16, 10, 12, 14, 25,
    ]
    aplicar_cabecera(ws, cabeceras, anchos)

    columnas_moneda = [8, 10, 12, 13, 16]
    fila = 2

    for fc in datos["facturas_cliente"]:
        signo = signo_factura(fc)
        lineas_fc = [
            l for l in datos["lineas_cliente"]
            if l["idfactura"] == fc["idfactura"]
        ]
        pct_iva = lineas_fc[0]["iva"] if lineas_fc else 0
        num_factura = fc.get("numero2") or fc.get("codigo", "")
        divisa = fc.get("_divisa_original", "EUR")
        tc = fc.get("_tc", 1)
        total_orig = fc.get("_total_original", fc["total"])

        ws.cell(row=fila, column=1, value=parsear_fecha(fc["fecha"]))
        ws.cell(row=fila, column=2, value=calcular_trimestre(fc["fecha"]))
        ws.cell(row=fila, column=3, value="Importacion limones")
        ws.cell(row=fila, column=4, value=num_factura)
        ws.cell(row=fila, column=5, value=fc.get("nombrecliente", ""))
        ws.cell(row=fila, column=6, value=fc.get("cifnif", ""))
        ws.cell(
            row=fila, column=7,
            value=concepto_factura(fc, datos["lineas_cliente"]),
        )
        ws.cell(row=fila, column=8, value=round(fc["neto"] * signo, 2))
        ws.cell(row=fila, column=9, value=pct_iva)
        ws.cell(row=fila, column=10, value=round(fc["totaliva"] * signo, 2))
        ws.cell(row=fila, column=11, value=0)
        ws.cell(row=fila, column=12, value=0)
        ws.cell(row=fila, column=13, value=round(fc["total"] * signo, 2))
        # Divisa/TC/Total Original justo despues de Total Factura
        ws.cell(row=fila, column=14, value=divisa)
        ws.cell(row=fila, column=15, value=tc if divisa != "EUR" else "")
        ws.cell(
            row=fila, column=16,
            value=round(total_orig * signo, 2) if divisa != "EUR" else "",
        )
        ws.cell(row=fila, column=17, value="Si" if fc.get("pagada") else "No")
        ws.cell(
            row=fila, column=18,
            value=parsear_fecha(fc["fecha"]) if fc.get("pagada") else "",
        )
        ws.cell(row=fila, column=19, value=fc.get("codpago", ""))
        ws.cell(row=fila, column=20, value=fc.get("observaciones", ""))

        formatear_moneda(ws, fila, columnas_moneda)
        fila += 1

    return ws


def crear_libro_gastos(wb, datos):
    """Pestana 2: Gastos (facturas de proveedor)."""
    ws = wb.create_sheet("Gastos")
    cabeceras = [
        "Fecha", "Trimestre", "Actividad", "N Factura", "Proveedor",
        "CIF Proveedor", "Concepto", "Base Imponible", "% IVA",
        "Cuota IVA", "IVA Deducible", "% IRPF", "Retencion IRPF",
        "Total Factura",
        "Divisa", "TC", "Total Original",
        "Pagada", "Fecha Pago", "Forma Pago", "Tipo Gasto",
        "Observaciones",
    ]
    anchos = [
        12, 10, 20, 16, 28, 18, 30, 16, 8, 14, 14, 8, 16, 14,
        8, 8, 16, 10, 12, 14, 20, 25,
    ]
    aplicar_cabecera(ws, cabeceras, anchos)

    columnas_moneda = [8, 10, 11, 13, 14, 17]
    fila = 2

    for fp in datos["facturas_proveedor"]:
        signo = signo_factura(fp)
        nombre_prov = fp.get("nombre", "")

        lineas_fp = [
            l for l in datos["lineas_proveedor"]
            if l["idfactura"] == fp["idfactura"]
        ]
        pct_iva = 0
        if lineas_fp:
            linea_mayor = max(lineas_fp, key=lambda l: abs(l["pvptotal"]))
            pct_iva = linea_mayor.get("iva", 0)

        iva_deducible = round(fp["totaliva"] * signo, 2)
        num_factura = fp.get("numproveedor") or fp.get("codigo", "")
        divisa = fp.get("_divisa_original", "EUR")
        tc = fp.get("_tc", 1)
        total_orig = fp.get("_total_original", fp["total"])

        ws.cell(row=fila, column=1, value=parsear_fecha(fp["fecha"]))
        ws.cell(row=fila, column=2, value=calcular_trimestre(fp["fecha"]))
        ws.cell(row=fila, column=3, value="Importacion limones")
        ws.cell(row=fila, column=4, value=num_factura)
        ws.cell(row=fila, column=5, value=nombre_prov)
        ws.cell(row=fila, column=6, value=fp.get("cifnif", ""))
        ws.cell(
            row=fila, column=7,
            value=concepto_factura(fp, datos["lineas_proveedor"]),
        )
        ws.cell(row=fila, column=8, value=round(fp["neto"] * signo, 2))
        ws.cell(row=fila, column=9, value=pct_iva)
        ws.cell(row=fila, column=10, value=round(fp["totaliva"] * signo, 2))
        ws.cell(row=fila, column=11, value=iva_deducible)
        ws.cell(row=fila, column=12, value=0)
        ws.cell(row=fila, column=13, value=0)
        ws.cell(row=fila, column=14, value=round(fp["total"] * signo, 2))
        # Divisa/TC/Total Original justo despues de Total Factura
        ws.cell(row=fila, column=15, value=divisa)
        ws.cell(row=fila, column=16, value=tc if divisa != "EUR" else "")
        ws.cell(
            row=fila, column=17,
            value=round(total_orig * signo, 2) if divisa != "EUR" else "",
        )
        ws.cell(row=fila, column=18, value="Si" if fp.get("pagada") else "No")
        ws.cell(
            row=fila, column=19,
            value=parsear_fecha(fp["fecha"]) if fp.get("pagada") else "",
        )
        ws.cell(row=fila, column=20, value=fp.get("codpago", ""))
        ws.cell(row=fila, column=21, value=clasificar_gasto(nombre_prov))
        ws.cell(row=fila, column=22, value=fp.get("observaciones", ""))

        formatear_moneda(ws, fila, columnas_moneda)
        fila += 1

    return ws


def crear_libro_bienes_inversion(wb):
    """Pestana 3: Bienes de inversion (vacia por ahora)."""
    ws = wb.create_sheet("Bienes Inversion")
    cabeceras = [
        "Fecha Adquisicion", "Actividad", "Descripcion", "Proveedor",
        "CIF Proveedor", "N Factura", "Base Imponible", "IVA",
        "Total", "Vida Util (anos)", "% Amortizacion",
        "Amortizacion Anual", "Amortizacion Acumulada",
        "Valor Neto Contable", "Fecha Baja", "Observaciones",
    ]
    anchos = [16, 14, 30, 25, 15, 14, 16, 14, 14, 16, 16, 18, 20, 18, 14, 25]
    aplicar_cabecera(ws, cabeceras, anchos)
    return ws


def desglosar_iva_lineas(lineas):
    """Desglosa bases e IVA por porcentaje a partir de lineas."""
    desglose = defaultdict(lambda: {"base": 0.0, "cuota": 0.0})
    for linea in lineas:
        pct = linea.get("iva", 0)
        base = linea.get("pvptotal", 0)
        desglose[pct]["base"] += base
        desglose[pct]["cuota"] += round(base * pct / 100, 2)
    return desglose


def crear_registro_facturas_emitidas(wb, datos):
    """Pestana 4: Registro de facturas emitidas con desglose por tipo IVA."""
    ws = wb.create_sheet("Registro Fact. Emitidas")
    cabeceras = [
        "Fecha Expedicion", "Fecha Operacion", "Trimestre", "Actividad",
        "N Factura", "Cliente", "CIF Cliente",
        "Base Imponible Exenta", "Base Imponible 21%", "Cuota IVA 21%",
        "Base Imponible 10%", "Cuota IVA 10%",
        "Base Imponible 4%", "Cuota IVA 4%",
        "Total Factura",
        "Divisa", "TC", "Total Original",
        "Observaciones",
    ]
    anchos = [
        16, 16, 10, 20, 16, 28, 15, 20, 18, 16, 18, 16, 18, 16, 14,
        8, 8, 16, 25,
    ]
    aplicar_cabecera(ws, cabeceras, anchos)

    columnas_moneda = [8, 9, 10, 11, 12, 13, 14, 15, 18]
    fila = 2

    for fc in datos["facturas_cliente"]:
        signo = signo_factura(fc)
        lineas_fc = [
            l for l in datos["lineas_cliente"]
            if l["idfactura"] == fc["idfactura"]
        ]
        desglose = desglosar_iva_lineas(lineas_fc)
        num_factura = fc.get("numero2") or fc.get("codigo", "")
        divisa = fc.get("_divisa_original", "EUR")
        tc = fc.get("_tc", 1)
        total_orig = fc.get("_total_original", fc["total"])

        ws.cell(row=fila, column=1, value=parsear_fecha(fc["fecha"]))
        ws.cell(row=fila, column=2, value=parsear_fecha(fc["fecha"]))
        ws.cell(row=fila, column=3, value=calcular_trimestre(fc["fecha"]))
        ws.cell(row=fila, column=4, value="Importacion limones")
        ws.cell(row=fila, column=5, value=num_factura)
        ws.cell(row=fila, column=6, value=fc.get("nombrecliente", ""))
        ws.cell(row=fila, column=7, value=fc.get("cifnif", ""))

        # Exentas (IVA 0%)
        ws.cell(
            row=fila, column=8,
            value=round(desglose[0]["base"] * signo, 2) if 0 in desglose else 0,
        )
        # 21%
        ws.cell(
            row=fila, column=9,
            value=round(desglose[21]["base"] * signo, 2) if 21 in desglose else 0,
        )
        ws.cell(
            row=fila, column=10,
            value=round(desglose[21]["cuota"] * signo, 2) if 21 in desglose else 0,
        )
        # 10%
        ws.cell(
            row=fila, column=11,
            value=round(desglose[10]["base"] * signo, 2) if 10 in desglose else 0,
        )
        ws.cell(
            row=fila, column=12,
            value=round(desglose[10]["cuota"] * signo, 2) if 10 in desglose else 0,
        )
        # 4%
        ws.cell(
            row=fila, column=13,
            value=round(desglose[4]["base"] * signo, 2) if 4 in desglose else 0,
        )
        ws.cell(
            row=fila, column=14,
            value=round(desglose[4]["cuota"] * signo, 2) if 4 in desglose else 0,
        )

        ws.cell(row=fila, column=15, value=round(fc["total"] * signo, 2))
        ws.cell(row=fila, column=16, value=divisa)
        ws.cell(row=fila, column=17, value=tc if divisa != "EUR" else "")
        ws.cell(
            row=fila, column=18,
            value=round(total_orig * signo, 2) if divisa != "EUR" else "",
        )
        ws.cell(row=fila, column=19, value=fc.get("observaciones", ""))

        formatear_moneda(ws, fila, columnas_moneda)
        fila += 1

    return ws


def crear_registro_facturas_recibidas(wb, datos):
    """Pestana 5: Registro de facturas recibidas con desglose por tipo IVA."""
    ws = wb.create_sheet("Registro Fact. Recibidas")
    cabeceras = [
        "Fecha Recepcion", "Fecha Factura", "Trimestre", "Actividad",
        "N Factura", "Proveedor", "CIF Proveedor",
        "Base Imponible 21%", "Cuota IVA 21%",
        "Base Imponible 10%", "Cuota IVA 10%",
        "Base Imponible 4%", "Cuota IVA 4%",
        "Base Exenta / IVA 0%",
        "IVA Deducible", "Total Factura",
        "Divisa", "TC", "Total Original",
        "Tipo Gasto", "Observaciones",
    ]
    anchos = [
        16, 14, 10, 20, 16, 28, 18, 18, 16, 18, 16, 18, 16, 20, 14, 14,
        8, 8, 16, 20, 25,
    ]
    aplicar_cabecera(ws, cabeceras, anchos)

    columnas_moneda = [8, 9, 10, 11, 12, 13, 14, 15, 16, 19]
    fila = 2

    for fp in datos["facturas_proveedor"]:
        signo = signo_factura(fp)
        lineas_fp = [
            l for l in datos["lineas_proveedor"]
            if l["idfactura"] == fp["idfactura"]
        ]
        desglose = desglosar_iva_lineas(lineas_fp)
        nombre_prov = fp.get("nombre", "")
        num_factura = fp.get("numproveedor") or fp.get("codigo", "")
        divisa = fp.get("_divisa_original", "EUR")
        tc = fp.get("_tc", 1)
        total_orig = fp.get("_total_original", fp["total"])

        ws.cell(row=fila, column=1, value=parsear_fecha(fp["fecha"]))
        ws.cell(row=fila, column=2, value=parsear_fecha(fp["fecha"]))
        ws.cell(row=fila, column=3, value=calcular_trimestre(fp["fecha"]))
        ws.cell(row=fila, column=4, value="Importacion limones")
        ws.cell(row=fila, column=5, value=num_factura)
        ws.cell(row=fila, column=6, value=nombre_prov)
        ws.cell(row=fila, column=7, value=fp.get("cifnif", ""))

        # 21%
        ws.cell(
            row=fila, column=8,
            value=round(desglose[21]["base"] * signo, 2) if 21 in desglose else 0,
        )
        ws.cell(
            row=fila, column=9,
            value=round(desglose[21]["cuota"] * signo, 2) if 21 in desglose else 0,
        )
        # 10%
        ws.cell(
            row=fila, column=10,
            value=round(desglose[10]["base"] * signo, 2) if 10 in desglose else 0,
        )
        ws.cell(
            row=fila, column=11,
            value=round(desglose[10]["cuota"] * signo, 2) if 10 in desglose else 0,
        )
        # 4%
        ws.cell(
            row=fila, column=12,
            value=round(desglose[4]["base"] * signo, 2) if 4 in desglose else 0,
        )
        ws.cell(
            row=fila, column=13,
            value=round(desglose[4]["cuota"] * signo, 2) if 4 in desglose else 0,
        )
        # Exentas / IVA 0%
        ws.cell(
            row=fila, column=14,
            value=round(desglose[0]["base"] * signo, 2) if 0 in desglose else 0,
        )

        # IVA deducible total
        ws.cell(row=fila, column=15, value=round(fp["totaliva"] * signo, 2))
        ws.cell(row=fila, column=16, value=round(fp["total"] * signo, 2))
        # Divisa/TC/Total Original
        ws.cell(row=fila, column=17, value=divisa)
        ws.cell(row=fila, column=18, value=tc if divisa != "EUR" else "")
        ws.cell(
            row=fila, column=19,
            value=round(total_orig * signo, 2) if divisa != "EUR" else "",
        )
        ws.cell(row=fila, column=20, value=clasificar_gasto(nombre_prov))
        ws.cell(row=fila, column=21, value=fp.get("observaciones", ""))

        formatear_moneda(ws, fila, columnas_moneda)
        fila += 1

    return ws


def crear_resumen_trimestral(wb, datos):
    """Pestana 6: Resumen trimestral con IVA repercutido/soportado."""
    ws = wb.create_sheet("Resumen Trimestral")
    cabeceras = [
        "Trimestre", "Actividad",
        "Ingresos Base", "IVA Repercutido",
        "Gastos Base", "IVA Soportado Deducible",
        "Resultado IVA (303)", "Observaciones",
    ]
    anchos = [12, 20, 16, 18, 16, 22, 18, 30]
    aplicar_cabecera(ws, cabeceras, anchos)

    columnas_moneda = [3, 4, 5, 6, 7]

    # Acumular por trimestre
    resumen = {}
    for t in ["T1", "T2", "T3", "T4"]:
        resumen[t] = {
            "ingresos_base": 0.0,
            "iva_repercutido": 0.0,
            "gastos_base": 0.0,
            "iva_soportado": 0.0,
        }

    for fc in datos["facturas_cliente"]:
        signo = signo_factura(fc)
        t = calcular_trimestre(fc["fecha"])
        resumen[t]["ingresos_base"] += fc["neto"] * signo
        resumen[t]["iva_repercutido"] += fc["totaliva"] * signo

    for fp in datos["facturas_proveedor"]:
        signo = signo_factura(fp)
        t = calcular_trimestre(fp["fecha"])
        resumen[t]["gastos_base"] += fp["neto"] * signo
        resumen[t]["iva_soportado"] += fp["totaliva"] * signo

    fila = 2
    for t in ["T1", "T2", "T3", "T4"]:
        r = resumen[t]
        resultado_iva = round(r["iva_repercutido"] - r["iva_soportado"], 2)

        ws.cell(row=fila, column=1, value=t)
        ws.cell(row=fila, column=2, value="Importacion limones")
        ws.cell(row=fila, column=3, value=round(r["ingresos_base"], 2))
        ws.cell(row=fila, column=4, value=round(r["iva_repercutido"], 2))
        ws.cell(row=fila, column=5, value=round(r["gastos_base"], 2))
        ws.cell(row=fila, column=6, value=round(r["iva_soportado"], 2))
        ws.cell(row=fila, column=7, value=resultado_iva)
        ws.cell(row=fila, column=8, value="")

        formatear_moneda(ws, fila, columnas_moneda)
        fila += 1

    return ws


def crear_conciliacion_bancaria(wb):
    """Pestana 7: Conciliacion bancaria (vacia por ahora)."""
    ws = wb.create_sheet("Conciliacion Bancaria")
    cabeceras = [
        "Fecha Movimiento", "Trimestre", "Concepto Banco",
        "Importe", "Tipo (Ingreso/Gasto)", "Actividad",
        "Factura Asociada", "Proveedor/Cliente",
        "Conciliado", "Observaciones",
    ]
    anchos = [16, 10, 30, 14, 20, 14, 18, 25, 12, 25]
    aplicar_cabecera(ws, cabeceras, anchos)
    return ws


def crear_diario_contable(wb, datos):
    """Pestana 8: Diario contable (asientos + partidas)."""
    ws = wb.create_sheet("Diario Contable")
    cabeceras = [
        "Fecha", "N Asiento", "Subcuenta", "Descripcion",
        "Debe", "Haber",
    ]
    anchos = [12, 12, 16, 50, 16, 16]
    aplicar_cabecera(ws, cabeceras, anchos)

    columnas_moneda = [5, 6]

    # Indexar partidas por idasiento
    partidas_por_asiento = defaultdict(list)
    for p in datos["partidas"]:
        partidas_por_asiento[p["idasiento"]].append(p)

    fila = 2
    for asiento in datos["asientos"]:
        id_asiento = asiento["idasiento"]
        partidas_asiento = sorted(
            partidas_por_asiento.get(id_asiento, []),
            key=lambda p: p.get("orden", 0),
        )
        for partida in partidas_asiento:
            cod_sub = partida.get("codsubcuenta", "")
            descripcion = datos["mapa_subcuentas"].get(cod_sub, "")
            concepto = partida.get("concepto", "")
            # Mostrar descripcion de subcuenta + concepto del asiento
            desc_completa = descripcion
            if concepto and concepto != descripcion:
                desc_completa = f"{descripcion} - {concepto}" if descripcion else concepto

            ws.cell(row=fila, column=1, value=parsear_fecha(asiento["fecha"]))
            ws.cell(row=fila, column=2, value=asiento.get("numero", id_asiento))
            ws.cell(row=fila, column=3, value=cod_sub)
            ws.cell(row=fila, column=4, value=desc_completa)
            ws.cell(row=fila, column=5, value=round(partida.get("debe", 0), 2))
            ws.cell(row=fila, column=6, value=round(partida.get("haber", 0), 2))

            formatear_moneda(ws, fila, columnas_moneda)
            fila += 1

    return ws


def crear_balance_sumas_saldos(wb, datos):
    """Pestana 9: Balance de sumas y saldos agrupado por subcuenta."""
    ws = wb.create_sheet("Balance Sumas y Saldos")
    cabeceras = [
        "Subcuenta", "Descripcion", "Total Debe", "Total Haber",
        "Saldo Deudor", "Saldo Acreedor",
    ]
    anchos = [16, 40, 16, 16, 16, 16]
    aplicar_cabecera(ws, cabeceras, anchos)

    columnas_moneda = [3, 4, 5, 6]

    # Agrupar partidas por subcuenta
    sumas = defaultdict(lambda: {"debe": 0.0, "haber": 0.0})
    for p in datos["partidas"]:
        cod = p.get("codsubcuenta", "")
        sumas[cod]["debe"] += p.get("debe", 0)
        sumas[cod]["haber"] += p.get("haber", 0)

    # Ordenar por codigo de subcuenta
    fila = 2
    for cod in sorted(sumas.keys()):
        total_debe = round(sumas[cod]["debe"], 2)
        total_haber = round(sumas[cod]["haber"], 2)
        saldo = round(total_debe - total_haber, 2)
        saldo_deudor = saldo if saldo > 0 else 0
        saldo_acreedor = abs(saldo) if saldo < 0 else 0

        descripcion = datos["mapa_subcuentas"].get(cod, "")

        ws.cell(row=fila, column=1, value=cod)
        ws.cell(row=fila, column=2, value=descripcion)
        ws.cell(row=fila, column=3, value=total_debe)
        ws.cell(row=fila, column=4, value=total_haber)
        ws.cell(row=fila, column=5, value=saldo_deudor)
        ws.cell(row=fila, column=6, value=saldo_acreedor)

        formatear_moneda(ws, fila, columnas_moneda)
        fila += 1

    # Fila de totales
    if fila > 2:
        total_fila = fila
        estilo_total = Font(bold=True, size=11)

        ws.cell(row=total_fila, column=1, value="TOTALES")
        ws.cell(row=total_fila, column=1).font = estilo_total

        for col in [3, 4, 5, 6]:
            letra = get_column_letter(col)
            formula = f"=SUM({letra}2:{letra}{fila - 1})"
            ws.cell(row=total_fila, column=col, value=formula)
            ws.cell(row=total_fila, column=col).font = estilo_total
            ws.cell(row=total_fila, column=col).number_format = estilo_moneda()

    return ws


def crear_validacion(wb, datos):
    """Pestana 10: Validacion cruzada de datos entre pestanas."""
    ws = wb.create_sheet("VALIDACION")
    from openpyxl.styles import PatternFill

    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    amarillo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    negrita = Font(bold=True, size=12)
    fmt = estilo_moneda()

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14

    fila = 1
    ws.cell(row=fila, column=1, value="VALIDACION CRUZADA DE DATOS").font = negrita
    fila += 2

    # --- Calculos desde datos ---
    # Facturas
    fc_total = sum(f["total"] * signo_factura(f) for f in datos["facturas_cliente"])
    fp_total = sum(f["total"] * signo_factura(f) for f in datos["facturas_proveedor"])
    fc_base = sum(f["neto"] * signo_factura(f) for f in datos["facturas_cliente"])
    fp_base = sum(f["neto"] * signo_factura(f) for f in datos["facturas_proveedor"])
    fc_iva = sum(f["totaliva"] * signo_factura(f) for f in datos["facturas_cliente"])
    fp_iva = sum(f["totaliva"] * signo_factura(f) for f in datos["facturas_proveedor"])

    # Partidas (diario)
    total_debe_diario = sum(float(p["debe"]) for p in datos["partidas"])
    total_haber_diario = sum(float(p["haber"]) for p in datos["partidas"])

    # Subcuentas — agrupar partidas
    saldos = defaultdict(lambda: {"debe": 0, "haber": 0})
    for p in datos["partidas"]:
        cod = p["codsubcuenta"]
        saldos[cod]["debe"] += float(p["debe"])
        saldos[cod]["haber"] += float(p["haber"])

    ventas_700 = saldos.get("7000000000", {}).get("haber", 0)
    # Saldo NETO de 600 (debe - haber) contempla NCs que van a HABER de 600
    compras_600_neto = (
        saldos.get("6000000000", {}).get("debe", 0)
        - saldos.get("6000000000", {}).get("haber", 0)
    )

    # IVA subcuentas (total incluyendo autoliquidacion intracomunitaria)
    iva_rep_total = sum(
        v["haber"] for k, v in saldos.items() if k.startswith("477")
    )
    iva_sop_total = sum(
        v["debe"] for k, v in saldos.items() if k.startswith("472")
    )

    # Autoliquidacion intracomunitaria:
    # Las facturas intracom se registran con IVA 0%, pero el asiento lleva
    # autoliquidacion (472 DEBE + 477 HABER). Este importe aparece en subcuentas
    # pero NO en las facturas. Lo identificamos via 4770000021 (solo autoliquidacion,
    # ya que no hay ventas al 21% — solo al 4% en 4770000004)
    autoliq_477 = saldos.get("4770000021", {}).get("haber", 0)
    autoliq_472 = autoliq_477  # Deberia ser igual en ambos lados

    # IVA PT reclasificado de 600 a 4709
    iva_pt = saldos.get("4709000000", {}).get("debe", 0)

    # Facturas en divisa extranjera
    n_usd_cli = sum(
        1 for f in datos["facturas_cliente"]
        if f.get("_divisa_original", "EUR") != "EUR"
    )
    n_usd_prov = sum(
        1 for f in datos["facturas_proveedor"]
        if f.get("_divisa_original", "EUR") != "EUR"
    )

    # --- Escribir validaciones ---
    def escribir_check(titulo, valor_a, valor_b, etiqueta_a, etiqueta_b,
                       nota=None, tolerancia=0.01):
        nonlocal fila
        ok = abs(valor_a - valor_b) < tolerancia
        ws.cell(row=fila, column=1, value=titulo).font = Font(bold=True)
        fila += 1
        ws.cell(row=fila, column=1, value=f"  {etiqueta_a}")
        ws.cell(row=fila, column=2, value=round(valor_a, 2))
        ws.cell(row=fila, column=2).number_format = fmt
        fila += 1
        ws.cell(row=fila, column=1, value=f"  {etiqueta_b}")
        ws.cell(row=fila, column=2, value=round(valor_b, 2))
        ws.cell(row=fila, column=2).number_format = fmt
        fila += 1
        diff = round(valor_a - valor_b, 2)
        ws.cell(row=fila, column=1, value=f"  Diferencia")
        ws.cell(row=fila, column=2, value=diff)
        ws.cell(row=fila, column=2).number_format = fmt
        estado = "OK" if ok else "ERROR"
        ws.cell(row=fila, column=3, value=estado)
        ws.cell(row=fila, column=3).fill = verde if ok else rojo
        if nota:
            fila += 1
            ws.cell(row=fila, column=1, value=f"  {nota}")
            ws.cell(row=fila, column=1).font = Font(italic=True, color="666666")
        fila += 2
        return ok

    resultados = []

    # 1. Cuadre diario
    resultados.append(escribir_check(
        "1. Cuadre Libro Diario (DEBE = HABER)",
        total_debe_diario, total_haber_diario,
        "Total DEBE", "Total HABER",
    ))

    # 2. Ventas: facturas vs subcuenta 700
    resultados.append(escribir_check(
        "2. Ingresos: Facturas emitidas vs Subcuenta 700",
        fc_base, ventas_700,
        "Base facturas cliente (EUR)", "Subcuenta 7000000000 HABER",
    ))

    # 3. IVA repercutido: facturas + autoliq intracom vs subcuenta 477
    resultados.append(escribir_check(
        "3. IVA Repercutido: Facturas + Autoliq. Intracom vs Subcuenta 477",
        fc_iva + autoliq_477, iva_rep_total,
        f"IVA facturas ({fc_iva:.2f}) + autoliq. intracom ({autoliq_477:.2f})",
        "Subcuentas 477x HABER",
        nota=f"Autoliquidacion intracomunitaria: {autoliq_477:.2f} EUR (Odoo+Transitainer)",
    ))

    # 4. Gastos: facturas vs subcuenta 600 NETO + 4709
    # Usamos NETO de 600 porque las NCs ponen importe en HABER de 600
    resultados.append(escribir_check(
        "4. Gastos: Facturas recibidas vs Subcuenta 600 neto + 4709",
        fp_base, compras_600_neto + iva_pt,
        "Base facturas proveedor (EUR, neto de NCs)",
        f"Subcuenta 600 neto ({compras_600_neto:.2f}) + 4709 ({iva_pt:.2f})",
        nota="600 neto = DEBE - HABER (contempla NCs). 4709 = IVA PT reclasificado",
    ))

    # 5. IVA soportado: facturas + autoliq intracom vs subcuenta 472
    resultados.append(escribir_check(
        "5. IVA Soportado: Facturas + Autoliq. Intracom vs Subcuenta 472",
        fp_iva + autoliq_472, iva_sop_total,
        f"IVA facturas ({fp_iva:.2f}) + autoliq. intracom ({autoliq_472:.2f})",
        "Subcuentas 472x DEBE",
        nota=f"Autoliquidacion intracomunitaria: {autoliq_472:.2f} EUR (Odoo+Transitainer)",
    ))

    # 6. Autoliquidacion equilibrada (472 == 477)
    autoliq_472_real = iva_sop_total - fp_iva
    autoliq_477_real = iva_rep_total - fc_iva
    resultados.append(escribir_check(
        "6. Autoliquidacion Intracom: Equilibrio 472 vs 477",
        autoliq_472_real, autoliq_477_real,
        "Exceso 472 sobre facturas prov",
        "Exceso 477 sobre facturas cli",
        nota="Ambos deben ser iguales (misma autoliquidacion en DEBE y HABER)",
    ))

    # --- Info divisas ---
    ws.cell(row=fila, column=1, value="INFO: Facturas en divisa extranjera").font = Font(bold=True)
    fila += 1
    ws.cell(row=fila, column=1, value=f"  Cliente: {n_usd_cli} facturas convertidas a EUR")
    fila += 1
    ws.cell(row=fila, column=1, value=f"  Proveedor: {n_usd_prov} facturas convertidas a EUR")
    fila += 1
    ws.cell(
        row=fila, column=1,
        value="  Ver columnas Divisa/TC/Total Original en pestanas Ingresos y Gastos",
    )
    fila += 2

    # --- Resumen ---
    ws.cell(row=fila, column=1, value="RESULTADO GLOBAL").font = negrita
    fila += 1
    n_ok = sum(resultados)
    n_total = len(resultados)
    if n_ok == n_total:
        ws.cell(row=fila, column=1, value=f"  {n_ok}/{n_total} validaciones OK")
        ws.cell(row=fila, column=1).fill = verde
    else:
        ws.cell(
            row=fila, column=1,
            value=f"  {n_ok}/{n_total} OK, {n_total - n_ok} ERRORES — revisar",
        )
        ws.cell(row=fila, column=1).fill = rojo

    return ws


# --- Main ---


def parsear_argumentos():
    """Parsea los argumentos de linea de comandos."""
    parser = argparse.ArgumentParser(
        description="Genera Excel de libros contables desde FacturaScripts"
    )
    parser.add_argument("ruta_salida", help="Ruta del archivo Excel de salida")
    parser.add_argument(
        "--empresa", type=int, default=1,
        help="ID de empresa en FacturaScripts (default: 1)",
    )
    return parser.parse_args()


def main():
    args = parsear_argumentos()

    # Crear directorio de salida si no existe
    directorio = os.path.dirname(args.ruta_salida)
    if directorio:
        os.makedirs(directorio, exist_ok=True)

    # Cargar datos de la API
    datos = cargar_datos(args.empresa)

    # Crear workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Generar pestanas
    crear_libro_ingresos(wb, datos)
    crear_libro_gastos(wb, datos)
    crear_libro_bienes_inversion(wb)
    crear_registro_facturas_emitidas(wb, datos)
    crear_registro_facturas_recibidas(wb, datos)
    crear_resumen_trimestral(wb, datos)
    crear_conciliacion_bancaria(wb)
    crear_diario_contable(wb, datos)
    crear_balance_sumas_saldos(wb, datos)
    crear_validacion(wb, datos)

    # Guardar
    wb.save(args.ruta_salida)
    print(f"Libros contables generados en: {args.ruta_salida}")


if __name__ == "__main__":
    main()
