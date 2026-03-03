# -*- coding: utf-8 -*-
"""
Genera gastos_gerardo_2025.xlsx para importar en Quipu.
Lee gastos_extraidos.json y para registros sin importe extrae el total del PDF.
"""

import pdfplumber
import json
import re
import os
import unicodedata
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

BASE_DIR = 'c:/Users/carli/PROYECTOS/CONTABILIDAD/clientes/gerardo-gonzalez-callejon/2025/batch1_originales'
JSON_PATH = 'c:/Users/carli/PROYECTOS/CONTABILIDAD/clientes/gerardo-gonzalez-callejon/2025/batch1_artefactos/gastos_extraidos.json'
TEMPLATE_PATH = 'c:/Users/carli/Downloads/gastos.xlsx'
SALIDA_PATH = 'c:/Users/carli/Downloads/gastos_gerardo_2025.xlsx'


def encontrar_pdf(carpeta, archivo):
    path = os.path.join(BASE_DIR, carpeta, archivo)
    if os.path.exists(path):
        return path
    carpeta_path = os.path.join(BASE_DIR, carpeta)
    if not os.path.exists(carpeta_path):
        return None
    nombre_norm = unicodedata.normalize('NFD', os.path.splitext(archivo)[0].lower()).encode('ascii', 'ignore').decode()
    for f in os.listdir(carpeta_path):
        if f.lower().endswith('.pdf'):
            f_norm = unicodedata.normalize('NFD', os.path.splitext(f)[0].lower()).encode('ascii', 'ignore').decode()
            if f_norm == nombre_norm:
                return os.path.join(carpeta_path, f)
    return None


def parsear_importe(s):
    s = s.strip()
    # Formato europeo: 1.234,56
    if re.match(r'^\d{1,3}(?:\.\d{3})*,\d{2}$', s):
        return float(s.replace('.', '').replace(',', '.'))
    # Simple con coma: 234,56
    if re.match(r'^\d+,\d{2}$', s):
        return float(s.replace(',', '.'))
    # Con punto decimal: 234.56
    if re.match(r'^\d+\.\d{2}$', s):
        return float(s)
    return None


def extraer_importe_pdf(path):
    """Extrae importe total del PDF bancario CaixaBank."""
    try:
        with pdfplumber.open(path) as pdf:
            texto = ' '.join(p.extract_text() or '' for p in pdf.pages)
    except Exception:
        return None

    if not texto.strip():
        return None

    def razonable(v):
        return v is not None and 0 < v < 50000

    # 1. Extracto movimiento: "Importe -X,XX euros"
    m = re.search(r'Importe\s+[-\u2013]\s*([\d.,]+)\s*euros', texto, re.IGNORECASE)
    if m:
        v = parsear_importe(m.group(1))
        if razonable(v):
            return v

    # 2. Domiciliacion SEPA: nombre + oficina + importe al final de linea
    m = re.search(r'GONZALEZ\s+CALLEJON\s+\S+\s+([\d.,]+)[\s\n]', texto, re.IGNORECASE)
    if m:
        v = parsear_importe(m.group(1))
        if razonable(v):
            return v

    # 3. Numero con simbolo euro
    for m in re.finditer(r'([\d.,]+)\s*(?:\u20ac|euros)', texto, re.IGNORECASE):
        v = parsear_importe(m.group(1))
        if razonable(v):
            return v

    # 4. Linea con Total o Importe
    for m in re.finditer(r'(?:Total|Importe)\s*:?\s*([\d.,]+)', texto, re.IGNORECASE):
        v = parsear_importe(m.group(1))
        if razonable(v):
            return v

    return None


def formatear_fecha(fecha_str):
    if not fecha_str or len(fecha_str) != 8:
        return None
    try:
        return datetime(int(fecha_str[:4]), int(fecha_str[4:6]), int(fecha_str[6:8]))
    except Exception:
        return None


def main():
    with open(JSON_PATH, encoding='utf-8') as f:
        datos = json.load(f)

    print(f'Total registros: {len(datos)}')

    # Filtrar no deducibles
    datos = [d for d in datos if not any(
        'NO deducible' in n or 'no deducible' in n
        for n in d.get('notas', [])
    )]

    # Ordenar por fecha
    datos.sort(key=lambda d: d.get('fecha', '') or '')

    # Recuperar importes de PDFs
    recuperados = 0
    for d in datos:
        if d.get('total') is not None or d.get('base') is not None:
            continue
        path = encontrar_pdf(d['carpeta'], d['archivo'])
        if not path:
            continue
        total = extraer_importe_pdf(path)
        if total:
            d['total'] = total
            recuperados += 1

    sin_datos_final = sum(1 for d in datos if d.get('total') is None and d.get('base') is None)
    print(f'Importes recuperados de PDF: {recuperados}')
    print(f'Sin importe (filas rojas): {sin_datos_final}')

    # Abrir template
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb['Sheet 1']

    # Limpiar datos existentes (fila 3 en adelante)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Colores por actividad
    fill_podologia  = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    fill_estetica   = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    fill_compartido = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    fill_sin_datos  = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')

    # Epigrafes IAE Gerardo
    # Podologia: Seccion II Profesionales, epigrafe 869
    # Estetica: Seccion I Empresarial, epigrafe 971.2
    actividad_map = {
        'podologia':  {'seccion': 'II',  'epigrafe': '869'},
        'estetica':   {'seccion': 'I',   'epigrafe': '971.2'},
        'compartido': {'seccion': '',    'epigrafe': ''},
    }

    fila = 3
    for d in datos:
        total     = d.get('total')
        base      = d.get('base')
        iva_pct   = d.get('iva_pct')
        irpf_pct  = d.get('irpf_pct')
        actividad = d.get('actividad', 'compartido')
        sin_datos = (total is None and base is None)

        if sin_datos:
            base_val = 0.0
        elif base is not None:
            base_val = base
        elif iva_pct:
            base_val = round(total / (1 + iva_pct / 100), 2)
        else:
            base_val = total

        tipo_doc    = 'Factura' if d.get('es_factura') else 'Factura Simplificada'
        fecha_dt    = formatear_fecha(d.get('fecha', ''))
        nif         = d.get('cif_proveedor') or None
        tipo_fiscal = 'NIF' if nif else 'none'
        nombre      = d.get('proveedor', '')
        concepto    = d.get('concepto', nombre)
        iae         = actividad_map.get(actividad, actividad_map['compartido'])

        if sin_datos:
            fill = fill_sin_datos
        elif actividad == 'podologia':
            fill = fill_podologia
        elif actividad == 'estetica':
            fill = fill_estetica
        else:
            fill = fill_compartido

        # 27 columnas segun plantilla Quipu
        valores = [
            tipo_doc,                                    # 1  tipo
            fecha_dt,                                    # 2  fecha emision
            None,                                        # 3  fecha vencimiento
            fecha_dt,                                    # 4  fecha pago
            'Pagado',                                    # 5  estado pago
            None,                                        # 6  numeracion
            nif,                                         # 7  NIF proveedor
            nombre,                                      # 8  nombre proveedor
            None,                                        # 9  correo
            None,                                        # 10 direccion
            None,                                        # 11 poblacion
            None,                                        # 12 CP
            'ES',                                        # 13 pais
            None,                                        # 14 telefono
            None,                                        # 15 num cuenta proveedor
            None,                                        # 16 swift
            None,                                        # 17 cod contable proveedor
            concepto,                                    # 18 concepto
            base_val,                                    # 19 base unitaria
            1,                                           # 20 cantidad
            iva_pct if iva_pct is not None else 0,       # 21 IVA%
            0,                                           # 22 recargo equivalencia
            irpf_pct if irpf_pct is not None else 0,     # 23 retencion%
            None,                                        # 24 num cuenta categoria
            iae['seccion'],                              # 25 seccion IAE
            iae['epigrafe'],                             # 26 epigrafe IAE
            tipo_fiscal,                                 # 27 tipo doc fiscal
        ]

        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=fila, column=col)
            cell.value = val
            cell.fill = fill
            if col in (2, 3, 4) and val is not None:
                cell.number_format = 'DD/MM/YYYY'

        fila += 1

    # Ajustar anchos
    anchos = [22, 12, 12, 12, 10, 15, 15, 30, 20, 25, 15, 8, 5, 12, 24, 12, 15, 40, 12, 8, 8, 8, 8, 15, 8, 10, 12]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.freeze_panes = 'A3'
    wb.save(SALIDA_PATH)

    suma = sum((d.get('total') or d.get('base') or 0) for d in datos)
    print(f'\nResumen:')
    print(f'  Registros escritos: {len(datos)}')
    print(f'  Con datos: {len(datos) - sin_datos_final}')
    print(f'  Sin importes (rojo): {sin_datos_final}')
    print(f'  Suma total importes: {suma:.2f} EUR')
    print(f'  Archivo: {SALIDA_PATH}')


if __name__ == '__main__':
    main()
