"""
Conciliacion Norma 43 vs Facturas OCR -- Gerardo Gonzalez Callejon
Tres cuentas CaixaBank 2025

Uso: python scripts/conciliar_gerardo_n43.py
Salida: Downloads/conciliacion_gerardo_2025.xlsx

Offsets Norma 43 (0-indexed, lineas de 80 chars):
  Record 11 (cabecera cuenta):
    [2:6]   banco        (4)
    [6:10]  oficina      (4)
    [10:12] digitos ctrl (2)
    [12:22] cuenta       (10)
    [20:26] fecha ini    (YYMMDD) -- NOTA: solapado con cuenta en este banco
    [26:32] fecha fin    (YYMMDD)
    [32]    signo saldo  (1=debe, 2=haber)
    [33:47] saldo ini    (14 digits, 2 dec)
    [49:80] nombre

  Record 22 (movimiento):
    [2:10]  banco+oficina (8, pueden ser espacios)
    [10:16] fecha op     (YYMMDD)
    [16:22] fecha valor  (YYMMDD)
    [22:24] concepto com (2)
    [24:26] concepto prop(2)
    [27]    D/C          (1=cargo, 2=abono) -- confirmado empiricamente
    [28:42] importe      (14 digits, 2 dec)
    [42:54] referencia 1 (12)
    [54:66] referencia 2 (12)
"""
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
import pandas as pd
from difflib import SequenceMatcher

N43_FILE   = Path("C:/Users/carli/Downloads/TT280226.423.txt")
EXCEL_FILE = Path("C:/Users/carli/Downloads/gastos_gerardo_2025.xlsx")
OUTPUT     = Path("C:/Users/carli/Downloads/conciliacion_gerardo_2025.xlsx")


# ------------------------------------------------------------------ #
# PASO 1 -- Parser Norma 43
# ------------------------------------------------------------------ #

def _fecha(s: str) -> date | None:
    try:
        return datetime.strptime("20" + s.strip(), "%Y%m%d").date()
    except Exception:
        return None


def _importe(s: str) -> float:
    try:
        return int(s.strip()) / 100
    except Exception:
        return 0.0


def parse_n43(filepath: Path):
    cuentas = []
    movimientos = []
    cuenta_actual = None
    mov_actual = None

    with open(filepath, "rb") as f:
        raw = f.read().decode("latin-1")

    for linea in raw.splitlines():
        if len(linea) < 2:
            continue
        tipo = linea[:2]

        # -- Cabecera cuenta --
        if tipo == "11":
            if mov_actual:
                movimientos.append(mov_actual)
                mov_actual = None

            banco   = linea[2:6].strip()
            oficina = linea[6:10].strip()
            cuenta  = linea[12:22].strip()
            fi      = _fecha(linea[20:26])
            ff      = _fecha(linea[26:32])
            dh      = linea[32]
            saldo   = _importe(linea[33:47])
            if dh == "1":
                saldo = -saldo
            nombre  = linea[49:80].strip()

            cuenta_actual = {
                "banco": banco, "oficina": oficina, "cuenta": cuenta,
                "fecha_ini": fi, "fecha_fin": ff,
                "saldo_ini": saldo, "nombre": nombre,
                "sufijo": cuenta[-4:] if len(cuenta) >= 4 else cuenta,
            }
            cuentas.append(cuenta_actual)

        # -- Movimiento --
        elif tipo == "22":
            if mov_actual:
                movimientos.append(mov_actual)

            fecha_op  = _fecha(linea[10:16])
            fecha_val = _fecha(linea[16:22])
            conc_c    = linea[22:24].strip()
            conc_p    = linea[24:26].strip()
            dc        = linea[27]              # '1'=cargo '2'=abono
            importe   = _importe(linea[28:42])
            ref1      = linea[42:54].strip()
            ref2      = linea[54:66].strip()

            if dc == "2":
                tipo_mov = "abono"
            else:
                tipo_mov = "cargo"
                importe  = -importe            # cargos en negativo

            mov_actual = {
                "sufijo":    cuenta_actual["sufijo"] if cuenta_actual else "",
                "cuenta":    cuenta_actual["cuenta"] if cuenta_actual else "",
                "fecha_op":  fecha_op,
                "fecha_val": fecha_val,
                "conc_c":    conc_c,
                "conc_p":    conc_p,
                "tipo":      tipo_mov,
                "importe":   importe,
                "ref1":      ref1,
                "ref2":      ref2,
                "desc":      "",
            }

        # -- Texto complementario --
        elif tipo == "23" and mov_actual is not None:
            txt = linea[4:80].strip()
            if txt:
                mov_actual["desc"] = (mov_actual["desc"] + " " + txt).strip()

        # -- Fin cuenta / fichero --
        elif tipo in ("33", "99"):
            if mov_actual:
                movimientos.append(mov_actual)
                mov_actual = None

    if mov_actual:
        movimientos.append(mov_actual)

    return cuentas, movimientos


# ------------------------------------------------------------------ #
# PASO 2 -- Cargar facturas Excel (formato Quipu / exportacion)
# ------------------------------------------------------------------ #

def cargar_facturas(filepath: Path) -> pd.DataFrame:
    # Fila 0 = grupos (Datos fiscales, Proveedor, Linea de factura...)
    # Fila 1 = cabeceras reales
    df = pd.read_excel(filepath, header=1)

    # Normalizar nombres de columna a ASCII lowercase
    nuevos = []
    seen = {}
    for c in df.columns:
        n = unicodedata.normalize("NFKD", str(c))
        n = n.encode("ascii", "ignore").decode("ascii").strip().lower()
        # Deduplicar
        if n in seen:
            seen[n] += 1
            n = f"{n}_{seen[n]}"
        else:
            seen[n] = 0
        nuevos.append(n)
    df.columns = nuevos

    # Columnas de fecha
    for col in ["fecha de emision", "fecha de pago", "fecha de vencimiento"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.date

    # Columnas numericas
    for col in ["base unitaria", "cantidad", "iva (%)", "retencion (%)"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Calcular importe neto a pagar
    qty = df.get("cantidad", pd.Series([1] * len(df))).fillna(1).replace(0, 1)
    iva = df.get("iva (%)", pd.Series([0] * len(df))).fillna(0)
    ret = df.get("retencion (%)", pd.Series([0] * len(df))).fillna(0)
    base = df.get("base unitaria", pd.Series([0] * len(df))).fillna(0)

    df["total_con_iva"] = base * qty * (1 + iva / 100)
    df["retencion_eur"] = base * qty * ret / 100
    df["pago_neto"]     = df["total_con_iva"] - df["retencion_eur"]

    return df


# ------------------------------------------------------------------ #
# PASO 3 -- Conciliacion
# ------------------------------------------------------------------ #

def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s).upper())
    s = s.encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _contiene(desc: str, nombre: str) -> bool:
    if not nombre or not desc:
        return False
    palabras = [p for p in _norm(nombre).split() if len(p) > 3]
    dn = _norm(desc)
    return any(p in dn for p in palabras)


def conciliar(movimientos: list[dict], facturas: pd.DataFrame) -> pd.DataFrame:
    DIAS_TOL = 7
    PCT_TOL  = 0.05     # 5% de diferencia en importe

    col_fecha = "fecha de pago" if "fecha de pago" in facturas.columns else "fecha de emision"
    col_nombre = "nombre" if "nombre" in facturas.columns else None
    col_num    = "numeracion" if "numeracion" in facturas.columns else None

    # Solo filas con fecha valida
    facts = facturas[facturas[col_fecha].notna()].to_dict("records")

    usadas = set()
    resultados = []

    # --- Emparejar cargos con facturas ---
    cargos = [m for m in movimientos if m["tipo"] == "cargo"]

    for mov in cargos:
        f_mov  = mov["fecha_op"]
        imp_abs = abs(mov["importe"])

        mejor_score = 0
        mejor_fac   = None
        mejor_idx   = None

        for idx, fac in enumerate(facts):
            if idx in usadas:
                continue

            f_fac = fac.get(col_fecha)
            if not isinstance(f_fac, date):
                continue

            diff_d = abs((f_mov - f_fac).days)
            if diff_d > DIAS_TOL:
                continue

            pago_neto = float(fac.get("pago_neto", 0) or 0)
            if pago_neto <= 0:
                continue

            diff_eur = abs(imp_abs - pago_neto)
            pct      = diff_eur / max(pago_neto, 0.01)

            if pct > PCT_TOL:
                continue

            # Score: 50pts fecha + 50pts importe + 20pts nombre en desc
            score = (
                max(0, (DIAS_TOL - diff_d) / DIAS_TOL) * 50
                + max(0, (PCT_TOL - pct) / PCT_TOL) * 50
            )
            nombre = str(fac.get(col_nombre, "")) if col_nombre else ""
            if _contiene(mov["desc"], nombre):
                score += 20

            if score > mejor_score:
                mejor_score = score
                mejor_fac   = fac
                mejor_idx   = idx

        estado = "CONCILIADO" if mejor_fac else "SIN FACTURA"
        pago   = round(float(mejor_fac.get("pago_neto", 0) or 0), 2) if mejor_fac else None

        resultados.append({
            "Cuenta":               mov["sufijo"],
            "Fecha operacion":      f_mov,
            "Tipo":                 "CARGO",
            "Importe banco":        round(-imp_abs, 2),
            "Descripcion banco":    mov["desc"][:80],
            "Estado":               estado,
            "Num factura":          mejor_fac.get(col_num, "") if mejor_fac else "",
            "Proveedor":            mejor_fac.get(col_nombre, "") if mejor_fac and col_nombre else "",
            "Fecha factura":        mejor_fac.get(col_fecha, "") if mejor_fac else "",
            "Importe factura":      pago,
            "Diferencia":           round(-imp_abs - (pago or 0), 2) if pago is not None else None,
            "Score":                round(mejor_score, 1),
        })

        if mejor_idx is not None:
            usadas.add(mejor_idx)

    # --- Facturas sin apunte bancario ---
    for idx, fac in enumerate(facts):
        if idx in usadas:
            continue
        pago = round(float(fac.get("pago_neto", 0) or 0), 2)
        resultados.append({
            "Cuenta":            "-",
            "Fecha operacion":   "",
            "Tipo":              "-",
            "Importe banco":     None,
            "Descripcion banco": "",
            "Estado":            "FACTURA SIN APUNTE",
            "Num factura":       fac.get(col_num, ""),
            "Proveedor":         fac.get(col_nombre, "") if col_nombre else "",
            "Fecha factura":     fac.get(col_fecha, ""),
            "Importe factura":   pago,
            "Diferencia":        None,
            "Score":             None,
        })

    # --- Abonos (ingresos) para informacion ---
    for mov in movimientos:
        if mov["tipo"] == "abono":
            resultados.append({
                "Cuenta":            mov["sufijo"],
                "Fecha operacion":   mov["fecha_op"],
                "Tipo":              "ABONO",
                "Importe banco":     round(mov["importe"], 2),
                "Descripcion banco": mov["desc"][:80],
                "Estado":            "INGRESO",
                "Num factura":       "",
                "Proveedor":         "",
                "Fecha factura":     "",
                "Importe factura":   None,
                "Diferencia":        None,
                "Score":             None,
            })

    return pd.DataFrame(resultados)


# ------------------------------------------------------------------ #
# PASO 4 -- Exportar Excel con formato
# ------------------------------------------------------------------ #

def exportar_excel(df: pd.DataFrame, cuentas: list, movimientos: list, filepath: Path):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    VERDE   = PatternFill("solid", fgColor="C6EFCE")
    ROJO    = PatternFill("solid", fgColor="FFC7CE")
    NARANJA = PatternFill("solid", fgColor="FFEB9C")
    AZUL    = PatternFill("solid", fgColor="BDD7EE")
    CAB_BG  = PatternFill("solid", fgColor="2F4F7F")
    CAB_FT  = Font(bold=True, color="FFFFFF")
    BD      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _cabecera(ws, cols):
        for ci, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=c)
            cell.fill = CAB_BG
            cell.font = CAB_FT
            cell.alignment = Alignment(horizontal="center")
            cell.border = BD

    # ---- Hoja 1: Conciliacion ----
    ws = wb.active
    ws.title = "Conciliacion"
    cols = list(df.columns)
    _cabecera(ws, cols)

    for ri, row in df.iterrows():
        estado = row.get("Estado", "")
        for ci, col in enumerate(cols, 1):
            val = row[col]
            if hasattr(val, "date"):
                val = val.date()
            elif val != val:   # NaN
                val = None
            cell = ws.cell(row=ri + 2, column=ci, value=val)
            cell.border = BD
            if estado == "CONCILIADO":
                cell.fill = VERDE
            elif estado == "SIN FACTURA":
                cell.fill = ROJO
            elif estado == "FACTURA SIN APUNTE":
                cell.fill = NARANJA
            elif estado == "INGRESO":
                cell.fill = AZUL

    anchos = {
        "Descripcion banco": 50, "Proveedor": 30, "Estado": 22,
        "Num factura": 16, "Cuenta": 8, "Tipo": 10,
    }
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = anchos.get(col, 16)
    ws.freeze_panes = "A2"

    # ---- Hoja 2: Movimientos brutos ----
    ws2 = wb.create_sheet("Movimientos N43")
    cab2 = ["Cuenta", "Fecha op", "Fecha valor", "Tipo", "Importe",
            "Conc C", "Conc P", "Ref1", "Ref2", "Descripcion"]
    _cabecera(ws2, cab2)
    CARGO_BG = PatternFill("solid", fgColor="FFF2CC")
    ABONO_BG = PatternFill("solid", fgColor="E2EFDA")
    for ri, m in enumerate(movimientos, 2):
        vals = [m["sufijo"], m["fecha_op"], m["fecha_val"], m["tipo"],
                round(m["importe"], 2), m["conc_c"], m["conc_p"],
                m["ref1"], m["ref2"], m["desc"][:80]]
        bg = CARGO_BG if m["tipo"] == "cargo" else ABONO_BG
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            c.border = BD
            c.fill   = bg
    ws2.column_dimensions["J"].width = 55
    ws2.freeze_panes = "A2"

    # ---- Hoja 3: Cuentas ----
    ws3 = wb.create_sheet("Cuentas")
    _cabecera(ws3, ["Banco", "Oficina", "Cuenta", "Titular", "Inicio", "Fin", "Saldo inicial"])
    for ri, ct in enumerate(cuentas, 2):
        ws3.append([ct["banco"], ct["oficina"], ct["cuenta"], ct["nombre"],
                    ct["fecha_ini"], ct["fecha_fin"], ct["saldo_ini"]])
    ws3.column_dimensions["D"].width = 35

    # ---- Hoja 4: Estadisticas ----
    ws4 = wb.create_sheet("Estadisticas")
    cargos_list  = [m for m in movimientos if m["tipo"] == "cargo"]
    abonos_list  = [m for m in movimientos if m["tipo"] == "abono"]
    conciliados  = df[df["Estado"] == "CONCILIADO"]
    sin_fac      = df[df["Estado"] == "SIN FACTURA"]
    sin_apunte   = df[df["Estado"] == "FACTURA SIN APUNTE"]

    stats = [
        ["CONCILIACION GERARDO GONZALEZ CALLEJON 2025", ""],
        ["", ""],
        ["--- MOVIMIENTOS BANCARIOS ---", ""],
        ["Total movimientos N43",       len(movimientos)],
        ["  Cargos (gastos/pagos)",     len(cargos_list)],
        ["  Total cargos (EUR)",         round(sum(abs(m["importe"]) for m in cargos_list), 2)],
        ["  Abonos (ingresos/cobros)",  len(abonos_list)],
        ["  Total abonos (EUR)",         round(sum(m["importe"] for m in abonos_list), 2)],
        ["", ""],
        ["--- FACTURAS OCR ---", ""],
        ["Total facturas",              len(conciliados) + len(sin_apunte)],
        ["", ""],
        ["--- CONCILIACION ---", ""],
        ["Cargos conciliados con factura",   len(conciliados)],
        ["Cargos SIN factura encontrada",    len(sin_fac)],
        ["Facturas SIN apunte bancario",     len(sin_apunte)],
        ["", ""],
        ["% Cargos conciliados",    f"{100*len(conciliados)/max(len(cargos_list),1):.1f}%"],
        ["% Facturas conciliadas",  f"{100*len(conciliados)/max(len(conciliados)+len(sin_apunte),1):.1f}%"],
    ]

    for row in stats:
        ws4.append(row)
    ws4.column_dimensions["A"].width = 40
    ws4.column_dimensions["B"].width = 20
    ws4["A1"].font = Font(bold=True, size=12)

    wb.save(filepath)


# ------------------------------------------------------------------ #
# MAIN
# ------------------------------------------------------------------ #

def main():
    print("=== Conciliacion Norma 43 vs OCR -- Gerardo Gonzalez ===")
    print()

    print("Parseando N43...")
    cuentas, movimientos = parse_n43(N43_FILE)

    print(f"  {len(cuentas)} cuentas:")
    for ct in cuentas:
        print(f"    {ct['banco']}-{ct['oficina']}-{ct['cuenta']} | "
              f"{ct['fecha_ini']} a {ct['fecha_fin']} | "
              f"Saldo ini: {ct['saldo_ini']:.2f} EUR")

    cargos_list  = [m for m in movimientos if m["tipo"] == "cargo"]
    abonos_list  = [m for m in movimientos if m["tipo"] == "abono"]
    print(f"  {len(movimientos)} movimientos: "
          f"{len(cargos_list)} cargos ({sum(abs(m['importe']) for m in cargos_list):.2f} EUR), "
          f"{len(abonos_list)} abonos ({sum(m['importe'] for m in abonos_list):.2f} EUR)")

    print()
    print("Cargando facturas Excel...")
    facturas = cargar_facturas(EXCEL_FILE)
    print(f"  {len(facturas)} lineas de facturas")
    col_fecha = "fecha de pago" if "fecha de pago" in facturas.columns else "fecha de emision"
    con_fecha = facturas[facturas[col_fecha].notna()]
    print(f"  {len(con_fecha)} facturas con fecha de pago")

    print()
    print("Conciliando...")
    resultado = conciliar(movimientos, facturas)

    conciliados = resultado[resultado["Estado"] == "CONCILIADO"]
    sin_fac     = resultado[resultado["Estado"] == "SIN FACTURA"]
    sin_apunte  = resultado[resultado["Estado"] == "FACTURA SIN APUNTE"]
    ingresos    = resultado[resultado["Estado"] == "INGRESO"]

    print(f"  OK  Conciliados con factura:    {len(conciliados)}")
    print(f"  XX  Cargos SIN factura:          {len(sin_fac)}")
    print(f"  --  Facturas SIN apunte bancario:{len(sin_apunte)}")
    print(f"  UP  Abonos/ingresos:             {len(ingresos)}")

    if not sin_fac.empty:
        print()
        print("  Cargos sin factura encontrada:")
        for _, r in sin_fac.iterrows():
            print(f"    {r['Fecha operacion']}  {r['Importe banco']:>10.2f} EUR  {r['Descripcion banco'][:55]}")

    if not sin_apunte.empty:
        print()
        print("  Facturas sin apunte bancario:")
        for _, r in sin_apunte.iterrows():
            imp = r["Importe factura"] or 0
            print(f"    {str(r['Fecha factura']):<12}  {imp:>10.2f} EUR  "
                  f"{str(r['Proveedor'])[:35]:<35}  ({r['Num factura']})")

    print()
    print("Generando Excel...")
    exportar_excel(resultado, cuentas, movimientos, OUTPUT)
    print(f"Guardado: {OUTPUT}")


if __name__ == "__main__":
    main()
